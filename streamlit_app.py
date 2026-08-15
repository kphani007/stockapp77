"""
StockMerit - NSE RSI Screener
------------------------------
Daily screen of NSE stocks by 14-day RSI, with a per-stock detail view
covering market cap, support/resistance and a financial health scorecard.

Deploy free on Streamlit Community Cloud (share.streamlit.io). Locally:
  pip install -r requirements.txt
  streamlit run streamlit_app.py
"""

from __future__ import annotations

import datetime as dt
import html as _html
import json
import uuid
import io
import re
import time
from collections import Counter
from urllib.parse import quote
import xml.etree.ElementTree as ET

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RSI_PERIOD = 14
FUND_LIMIT = 100
MCAP_LARGE_CR, MCAP_MID_CR, MCAP_SMALL_CR = 20000, 5000, 500

NIFTY50 = [
    "ADANIENT.NS", "ADANIPORTS.NS", "APOLLOHOSP.NS", "ASIANPAINT.NS",
    "AXISBANK.NS", "BAJAJ-AUTO.NS", "BAJFINANCE.NS", "BAJAJFINSV.NS",
    "BEL.NS", "BHARTIARTL.NS", "CIPLA.NS", "COALINDIA.NS", "DRREDDY.NS",
    "EICHERMOT.NS", "GRASIM.NS", "HCLTECH.NS", "HDFCBANK.NS",
    "HDFCLIFE.NS", "HEROMOTOCO.NS", "HINDALCO.NS", "HINDUNILVR.NS",
    "ICICIBANK.NS", "INDUSINDBK.NS", "INFY.NS", "ITC.NS", "JSWSTEEL.NS",
    "KOTAKBANK.NS", "LT.NS", "M&M.NS", "MARUTI.NS", "NESTLEIND.NS",
    "NTPC.NS", "ONGC.NS", "POWERGRID.NS", "RELIANCE.NS", "SBILIFE.NS",
    "SBIN.NS", "SHRIRAMFIN.NS", "SUNPHARMA.NS", "TATACONSUM.NS",
    "TATAMOTORS.NS", "TATASTEEL.NS", "TCS.NS", "TECHM.NS", "TITAN.NS",
    "TRENT.NS", "ULTRACEMCO.NS", "WIPRO.NS",
]

HEADERS = [
    "Date", "Stock Symbol", "Sector", "Current Price (Rs)", "Volume",
    "RSI", "VWAP", "PE", "Sec PE", "Buy/Sell", "52 Week High (Rs)", "1 Year Target (Rs)",
]

COL_LABELS = {
    "Date": "Date", "Stock Symbol": "Stock Symbol", "Sector": "Sector",
    "Current Price (Rs)": "Current Price", "Volume": "Volume", "RSI": "RSI",
    "VWAP": "VWAP", "PE": "PE", "Sec PE": "Sec PE",
    "Buy/Sell": "Buy/Sell", "52 Week High (Rs)": "52 Week High",
    "1 Year Target (Rs)": "1 Year Forecast",
}

OI_SYMBOLS = [
    "RELIANCE", "SBIN", "ICICIBANK", "BAJFINANCE", "HDFCBANK",
    "TATAMOTORS", "MARUTI", "INFY", "TATASTEEL", "AXISBANK",
]
DHAN_QUOTE_URL = "https://api.dhan.co/v2/marketfeed/quote"
DHAN_HIST_URL = "https://api.dhan.co/v2/charts/historical"
DHAN_SCRIP_URL = "https://images.dhan.co/api-data/api-scrip-master-detailed.csv"


def _dhan_creds() -> tuple[str, str]:
    try:
        return (str(st.secrets.get("DHAN_CLIENT_ID", "")).strip(),
                str(st.secrets.get("DHAN_ACCESS_TOKEN", "")).strip())
    except Exception:
        return "", ""


@st.cache_data(ttl=60, show_spinner=False)
def _yf_oi_prices(symbols: tuple[str, ...]) -> list[dict]:
    """Live LTP + price change from yfinance (no auth). OI needs a Dhan feed,
    so it is returned as None here and shown as '-' rather than a stale guess."""
    rows = []
    for sym in symbols:
        try:
            tk = yf.Ticker(f"{sym}.NS")
            ltp = prev = None
            try:
                fi = tk.fast_info
                ltp = fi.get("last_price") if hasattr(fi, "get") else getattr(fi, "last_price", None)
                prev = fi.get("previous_close") if hasattr(fi, "get") else getattr(fi, "previous_close", None)
            except Exception:
                pass
            if ltp is None or prev is None:
                h = tk.history(period="5d")["Close"].dropna()
                if len(h) >= 2:
                    ltp, prev = float(h.iloc[-1]), float(h.iloc[-2])
            if ltp is None or not prev:
                continue
            rows.append({"sym": sym, "ltp": float(ltp),
                         "price_chg": (float(ltp) / float(prev) - 1) * 100,
                         "oi": None, "oi_chg": None})
        except Exception:
            continue
    rows.sort(key=lambda r: r["price_chg"], reverse=True)
    return rows


@st.cache_data(ttl=86400, show_spinner=False)
def _dhan_fut_map(symbols: tuple[str, ...]) -> dict:
    """Map each stock symbol -> nearest NSE stock-future security id (from scrip master)."""
    try:
        df = pd.read_csv(DHAN_SCRIP_URL, low_memory=False)
    except Exception:
        return {}
    up = {c.upper(): c for c in df.columns}

    def col(*names):
        for n in names:
            if n in up:
                return up[n]
        return None

    c_exch = col("SEM_EXM_EXCH_ID", "EXCH_ID")
    c_instr = col("SEM_INSTRUMENT_NAME", "INSTRUMENT", "SEM_EXCH_INSTRUMENT_TYPE")
    c_sec = col("SEM_SMST_SECURITY_ID", "SECURITY_ID")
    c_exp = col("SEM_EXPIRY_DATE", "EXPIRY")
    c_und = col("SM_SYMBOL_NAME", "SEM_UNDERLYING_SYMBOL", "UNDERLYING_SYMBOL",
                "SEM_TRADING_SYMBOL")
    if not all([c_exch, c_instr, c_sec, c_und]):
        return {}
    try:
        d = df[(df[c_exch].astype(str) == "NSE")
               & (df[c_instr].astype(str).str.contains("FUTSTK", case=False, na=False))]
        out = {}
        for s in symbols:
            rows = d[d[c_und].astype(str).str.upper() == s]
            if rows.empty:
                continue
            if c_exp:
                rows = rows.sort_values(c_exp)
            out[s] = int(rows.iloc[0][c_sec])
        return out
    except Exception:
        return {}


@st.cache_data(ttl=60, show_spinner=False)
def get_oi_buildup() -> tuple[list[dict], str]:
    """Return (rows, source). source: 'dhan' = live price+OI, 'yf' = live price
    only (no OI feed), 'none' = nothing reachable. Never returns invented data."""
    cid, tok = _dhan_creds()
    if not cid or not tok:
        rows = _yf_oi_prices(tuple(OI_SYMBOLS))
        return rows, ("yf" if rows else "none")
    try:
        secmap = _dhan_fut_map(tuple(OI_SYMBOLS))
        if not secmap:
            rows = _yf_oi_prices(tuple(OI_SYMBOLS))
            return rows, ("yf" if rows else "none")
        headers = {"access-token": tok, "client-id": cid,
                   "Content-Type": "application/json", "Accept": "application/json"}
        ids = list(secmap.values())
        q = requests.post(DHAN_QUOTE_URL, json={"NSE_FNO": ids},
                          headers=headers, timeout=8).json()
        qmap = (q.get("data") or {}).get("NSE_FNO") or {}
        today = dt.date.today()
        frm = (today - dt.timedelta(days=12)).isoformat()
        rows = []
        for sym, sid in secmap.items():
            qd = qmap.get(str(sid)) or {}
            ltp = qd.get("last_price", qd.get("ltp"))
            oi = qd.get("oi")
            hb = requests.post(DHAN_HIST_URL, headers=headers, timeout=8, json={
                "securityId": str(sid), "exchangeSegment": "NSE_FNO",
                "instrument": "FUTSTK", "expiryCode": 0,
                "fromDate": frm, "toDate": today.isoformat()}).json()
            closes = hb.get("close") or []
            ois = hb.get("open_interest") or hb.get("openInterest") or []
            if ltp is None and closes:
                ltp = closes[-1]
            if oi is None and ois:
                oi = ois[-1]
            prev_close = closes[-2] if len(closes) >= 2 else None
            prev_oi = ois[-2] if len(ois) >= 2 else None
            if None in (ltp, oi, prev_close, prev_oi) or not prev_close or not prev_oi:
                continue
            rows.append({"sym": sym, "ltp": float(ltp),
                         "price_chg": (float(ltp) / float(prev_close) - 1) * 100,
                         "oi": int(oi),
                         "oi_chg": (float(oi) / float(prev_oi) - 1) * 100})
        if not rows:
            _rows = _yf_oi_prices(tuple(OI_SYMBOLS))
            return _rows, ("yf" if _rows else "none")
        rows.sort(key=lambda r: r["oi_chg"], reverse=True)
        return rows, "dhan"
    except Exception:
        rows = _yf_oi_prices(tuple(OI_SYMBOLS))
        return rows, ("yf" if rows else "none")


def oi_signal(price_chg: float, oi_chg) -> tuple[str, str]:
    if oi_chg is None:
        return ("-", "#9AA7B2")
    if oi_chg >= 0:
        return ("Long Buildup", "#0B7A4B") if price_chg >= 0 else ("Short Buildup", "#B3261E")
    return ("Short Covering", "#0E7C86") if price_chg >= 0 else ("Long Unwinding", "#C77A0B")

UNIVERSES = ["All NIFTY Stocks", "NIFTY 50", "NIFTY 500"]

HEALTH_GROUPS = [
    ("Cash flow", ["Cash flow backs reported profit", "Operating cash flow positive",
                   "Free cash flow positive"]),
    ("Debt & solvency", ["Debt to equity", "Debt reduced vs last year",
                         "Interest coverage", "Current ratio"]),
    ("Efficiency", ["Receivables in line with sales", "Inventory in line with sales"]),
    ("Growth", ["Revenue growing (3y)", "Profit growing (3y)",
                "Profitable every year (3y)"]),
    ("Profitability", ["Return on equity", "Operating margin holding up",
                       "Net margin positive"]),
]

RECO_INFO = {
    "Strong Buy": "Highest-conviction analyst call — consensus strongly expects the "
                  "price to rise over the next 12 months.",
    "Buy": "Analysts expect the price to rise over the next 12 months.",
    "Hold": "Analysts expect roughly flat performance — wait and watch.",
    "Sell": "Analysts expect the price to fall — consider trimming or exiting.",
}

_EQUITY_LIST_URL = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
_INDEX500_LIST_URL = "https://niftyindices.com/IndexConstituent/ind_nifty500list.csv"

SECTOR_PALETTE = ["#0E7C86", "#6B5BC7", "#C77A0B", "#1F6FB2",
                  "#0B7A4B", "#B3467A", "#2F8F5B", "#8A5A2B"]

DISCLAIMER = (
    "Everything shown here is for reference and general information only. It is not "
    "investment advice, not a recommendation to buy or sell any security, and not a "
    "solicitation of any kind. The Buy/Sell rating and 1-year target are third-party "
    "analyst consensus figures reproduced as-is, not the view of this app. Technical "
    "levels and health checks are mechanically computed from reported financials and "
    "may be wrong, incomplete or stale. Do your own research and consult a "
    "SEBI-registered adviser before investing. Securities investments carry market "
    "risk, including possible loss of capital."
)

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Archivo:wght@500;600;700&family=Inter:wght@400;500&family=IBM+Plex+Mono:wght@400;500;600&display=swap');

:root{
  --ink:#111C2B; --panel:#F5F8FA; --line:#DDE6ED;
  --teal:#0E7C86; --violet:#6B5BC7; --amber:#C77A0B; --blue:#1F6FB2;
  --green:#0B7A4B; --coral:#D8542F; --red:#B3261E; --plum:#B3467A;
  --text:#16202B; --muted:#5E6E7E;
}
html, body, [class*="css"], .stMarkdown { font-family:'Inter', sans-serif; color:var(--text); }
h1,h2,h3,h4,h5 { font-family:'Archivo', sans-serif; letter-spacing:-0.02em; }
[data-testid="stDataFrame"] { font-family:'IBM Plex Mono', monospace; }

.band{
  position:sticky; top:0; z-index:1000;
  background:linear-gradient(100deg,#111C2B 0%,#123A44 55%,#0E7C86 100%);
  border-radius:12px; padding:7px 16px; margin-bottom:8px;
  display:flex; align-items:center; justify-content:space-between;
  flex-wrap:nowrap; gap:14px; box-shadow:0 8px 22px -16px rgba(17,28,43,0.6);
}
.band-left{ display:flex; align-items:baseline; gap:14px; min-width:0; flex-wrap:wrap; }
.band-name{ font-family:'Archivo',sans-serif; font-weight:700; font-size:1.15rem;
  color:#FFF; letter-spacing:-0.03em; line-height:1.1; display:flex; align-items:center; gap:8px; }
.band-name .brandmark{ width:22px; height:22px; border-radius:6px; background:#FFF;
  display:inline-flex; align-items:center; justify-content:center; }
.band-name .m{ color:#5CA8FF; }
.band-sub{ font-family:'IBM Plex Mono',monospace; font-size:0.74rem;
  color:#B6D8DC; margin:0; white-space:nowrap; }
.idxcard{ background:#FFF; border:1px solid #DDE6ED; border-radius:10px; padding:8px 11px; }
.idx-n{ font-size:0.66rem; text-transform:uppercase; letter-spacing:0.06em; color:#5E6E7E; font-weight:600; }
.idx-v{ font-family:'IBM Plex Mono',monospace; font-size:0.9rem; font-weight:600; margin-top:3px; }
.idx-c{ font-family:'IBM Plex Mono',monospace; font-size:0.74rem; font-weight:600; margin-top:2px; }
.band-date{ font-family:'IBM Plex Mono',monospace; font-size:0.72rem; color:#EAF6F7;
  background:rgba(255,255,255,0.14); border-radius:999px; padding:4px 12px;
  white-space:nowrap; }
.hdr-search div[data-baseweb="select"] > div{ min-height:38px; }

.sec-label{ font-size:0.72rem; text-transform:uppercase; letter-spacing:0.09em;
  color:var(--muted); font-weight:600; margin:2px 0 6px; }

.stat-row{ display:flex; flex-wrap:wrap; gap:10px; margin-bottom:4px; }
.stat{ flex:1 1 150px; border-radius:12px; padding:11px 13px; border:1px solid var(--line);
  border-top:4px solid var(--teal); background:#FFF; }
div[role="dialog"]{ max-width:600px !important; }
.stat-row .stat{ flex:1 1 120px; padding:9px 11px; }
.stat-row .stat-v{ font-size:1.0rem; }
.stat.v{ border-top-color:var(--violet); }
.stat.a{ border-top-color:var(--amber); }
.stat.b{ border-top-color:var(--blue); }
.stat.g{ border-top-color:var(--green); }
.stat-k{ font-size:0.68rem; text-transform:uppercase; letter-spacing:0.08em;
  color:var(--muted); font-weight:600; }
.stat-v{ font-family:'IBM Plex Mono',monospace; font-size:1.1rem; font-weight:600;
  margin-top:3px; line-height:1.2; }
.stat-n{ font-size:0.73rem; color:var(--muted); margin-top:3px; }

.score-row{ display:flex; gap:10px; flex-wrap:wrap; margin:6px 0 16px; }
.score{ flex:1 1 118px; border-radius:10px; padding:10px 13px; color:#FFF; }
.score.pos{ background:var(--green); }
.score.neu{ background:#6B7A8F; }
.score.neg{ background:var(--red); }
.score.nod{ background:#A8B4BF; }
.score-n{ font-family:'IBM Plex Mono',monospace; font-size:1.4rem; font-weight:600;
  line-height:1; }
.score-l{ font-size:0.75rem; margin-top:5px; opacity:0.92; }

.chk{ display:flex; align-items:baseline; gap:10px; padding:8px 0;
  border-bottom:1px solid var(--line); font-size:0.87rem; }
.chk:last-child{ border-bottom:none; }
.tag{ font-family:'IBM Plex Mono',monospace; font-size:0.65rem; font-weight:600;
  padding:3px 8px; border-radius:5px; text-transform:uppercase; letter-spacing:0.05em;
  flex:0 0 auto; min-width:64px; text-align:center; color:#FFF; }
.tag.Yes{ background:var(--green); }
.tag.No{ background:var(--red); }
.tag.Neutral{ background:#6B7A8F; }
.tag.NoData{ background:#BFC9D2; }
.chk-n{ flex:1 1 auto; }
.chk-d{ font-family:'IBM Plex Mono',monospace; font-size:0.78rem; color:var(--muted); }

.lvl{ display:flex; justify-content:space-between; padding:7px 11px; border-radius:7px;
  font-family:'IBM Plex Mono',monospace; font-size:0.85rem; margin-bottom:4px;
  color:#FFF; }
.lvl.r{ background:#D8542F; }
.lvl.r2{ background:#E07E5F; }
.lvl.p{ background:var(--violet); font-weight:600; }
.lvl.s{ background:#2F8F5B; }
.lvl.s2{ background:#5CAE82; }
.lvl.now{ background:var(--ink); font-weight:600; }

.disc{ font-size:0.76rem; color:var(--muted); line-height:1.6;
  border-top:1px solid var(--line); padding-top:14px; margin-top:10px; }
.stButton>button{ border-radius:9px; font-weight:500; padding:6px 10px; font-size:0.9rem; }
.stDownloadButton>button{ background:transparent; color:var(--teal);
  border:1px solid var(--teal); font-weight:600; border-radius:9px; padding:5px 14px; }
.stDownloadButton>button:hover{ background:#E3F2F3; color:var(--teal); }
div[role="dialog"] button[data-baseweb="tab"]{ font-family:'Archivo',sans-serif;
  font-weight:800; font-size:0.98rem; background:#EAF2FB; border-radius:8px 8px 0 0;
  padding:7px 20px; margin-right:6px; color:#1F6FB2; }
div[role="dialog"] button[data-baseweb="tab"][aria-selected="true"]{
  background:var(--teal); color:#FFF; }
div[role="dialog"] div[data-baseweb="tab-highlight"]{ display:none; }
div[role="dialog"] div[data-baseweb="tab-border"]{ display:none; }
div[role="dialog"] .stat{ padding:8px 10px; }
div[role="dialog"] .stat-v{ font-size:0.95rem; }
div[role="dialog"] .lvl{ padding:5px 10px; font-size:0.8rem; margin-bottom:3px; }
div[role="dialog"] .sec-label{ margin:0 0 4px; }

/* --- custom results table (matches the prototype) --- */
.sh-tablewrap{ border:1px solid #E5ECF1; border-radius:12px; overflow-x:auto; margin-top:6px; }
.sh-table{ width:100%; border-collapse:collapse; }
.sh-tablewrap{ max-height:72vh; overflow:auto; }
.sh-table thead th{ background:#F5F8FA; padding:11px 14px; font-size:0.68rem;
  text-transform:uppercase; letter-spacing:0.07em; color:var(--muted); font-weight:600;
  white-space:nowrap; position:sticky; top:0; z-index:3;
  box-shadow:inset 0 -1px 0 #DDE6ED; }
.sh-table .th-r{ text-align:right; } .sh-table .th-c{ text-align:center; }
.sh-table .th-l{ text-align:left; }
.sh-table td{ padding:12px 14px; border-top:1px solid #EEF3F6; white-space:nowrap;
  font-size:0.85rem; color:var(--text); }
.sh-table tbody tr{ transition:background .1s; }
.sh-table tbody tr:hover{ background:#F5F8FA; }
.sh-table .c-date{ font-family:'IBM Plex Mono',monospace; font-size:0.8rem; color:var(--muted); }
.sh-table a.c-sym{ font-family:'IBM Plex Mono',monospace; font-weight:600; font-size:0.86rem;
  color:var(--teal); text-decoration:none; }
.sh-table a.c-sym:hover{ text-decoration:underline; }
.sh-table .c-price{ text-decoration:underline dotted; text-underline-offset:3px;
  text-decoration-color:#9BB4C4; cursor:help; }
.sh-table .c-sec{ font-weight:500; }
.sh-table .c-num{ font-family:'IBM Plex Mono',monospace; text-align:right; }
.sh-table .c-muted{ color:var(--muted); }
.sh-table .c-rsi{ font-weight:600; font-size:0.9rem; }
.sh-table .c-vwap{ cursor:help; }
.stat{ resize:both; overflow:auto; min-width:120px; min-height:70px; }
.sh-table .c-reco{ text-align:center; }
.sh-table .pill{ font-family:'IBM Plex Mono',monospace; font-size:0.68rem; font-weight:600;
  padding:3px 9px; border-radius:6px; color:#FFF; white-space:nowrap; }
.sh-hint{ margin-top:10px; font-size:0.75rem; color:#8794A1; }

/* pull content up + strip default header/sidebar chrome */
[data-testid="stHeader"]{ background:transparent; height:0; }
.block-container{ padding-top:1.1rem !important; }
/* --- left nav (collapsible sidebar) --- */
[data-testid="stSidebar"]{ background:#111C2B; border-right:1px solid #1E2E42; }
[data-testid="stSidebar"] *{ color:#EAF6F7; }
[data-testid="stSidebar"] .nav-h{ font-family:'Archivo',sans-serif; font-size:0.68rem;
  letter-spacing:0.1em; text-transform:uppercase; color:#7FA8B5; font-weight:700;
  margin:2px 0 10px; }
[data-testid="stSidebar"] div[role="radiogroup"]{ display:flex; flex-direction:column; gap:3px; }
[data-testid="stSidebar"] div[role="radiogroup"] label{ padding:9px 12px; border-radius:9px;
  font-size:0.93rem; font-weight:500; cursor:pointer; width:100%; }
[data-testid="stSidebar"] div[role="radiogroup"] label > div:first-child{ display:none; }
[data-testid="stSidebar"] div[role="radiogroup"] label:hover{ background:rgba(255,255,255,0.09); }
[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked){ background:#0E7C86; }
[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) p{ font-weight:700; }
[data-testid="stSidebar"] .nav-foot{ font-size:0.7rem; color:#7FA8B5; line-height:1.5;
  margin-top:14px; }
/* movable + resizable detail dialog */
div[role="dialog"]{ resize:both; overflow:auto; min-width:340px; min-height:220px; }
div[role="dialog"]:hover{ box-shadow:0 20px 60px rgba(17,28,43,0.35); }
</style>
"""


# ------------------------------ indicators --------------------------------

def compute_rsi(close: pd.Series, period: int = RSI_PERIOD) -> float | None:
    close = close.dropna()
    if len(close) < period + 1:
        return None
    delta = close.diff()
    gain = delta.clip(lower=0.0)
    loss = -delta.clip(upper=0.0)
    avg_gain = gain.ewm(alpha=1.0 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1.0 / period, min_periods=period, adjust=False).mean()
    g, l = avg_gain.iloc[-1], avg_loss.iloc[-1]
    if pd.isna(g) or pd.isna(l):
        return None
    if l == 0:
        return 100.0
    return float(100.0 - (100.0 / (1.0 + (g / l))))


def classify_market_cap(mcap_inr: float | None) -> str:
    if not mcap_inr or mcap_inr <= 0:
        return "n/a"
    cr = mcap_inr / 1e7
    if cr >= MCAP_LARGE_CR:
        return "Large cap"
    if cr >= MCAP_MID_CR:
        return "Mid cap"
    if cr >= MCAP_SMALL_CR:
        return "Small cap"
    return "Micro cap"


def pivot_levels(h: float, l: float, c: float) -> dict:
    p = (h + l + c) / 3.0
    return {"R3": h + 2 * (p - l), "R2": p + (h - l), "R1": 2 * p - l, "Pivot": p,
            "S1": 2 * p - h, "S2": p - (h - l), "S3": l - 2 * (h - p)}


def swing_levels(close: pd.Series, price: float, window: int = 10):
    s = close.dropna()
    highs = s[(s.shift(window) < s) & (s.shift(-window) < s)]
    lows = s[(s.shift(window) > s) & (s.shift(-window) > s)]
    sup, res = lows[lows < price], highs[highs > price]
    return (float(sup.max()) if len(sup) else None,
            float(res.min()) if len(res) else None)


def sector_color(name) -> str:
    if not isinstance(name, str) or name in ("n/a", "not loaded", ""):
        return "#8794A1"
    return SECTOR_PALETTE[sum(ord(c) for c in name) % len(SECTOR_PALETTE)]


def compute_vwap(close: pd.Series, vol: pd.Series, n: int = 20):
    """20-day volume-weighted average price. Returns None if unavailable."""
    if close is None or vol is None:
        return None
    df = pd.concat([close, vol], axis=1).dropna()
    if df.empty:
        return None
    df = df.tail(n)
    tot_vol = float(df.iloc[:, 1].sum())
    if not tot_vol:
        return None
    return float((df.iloc[:, 0] * df.iloc[:, 1]).sum() / tot_vol)


def rsi_color(v) -> str:
    if not isinstance(v, (int, float)):
        return "#5E6E7E"
    if v >= 80:
        return "#B3261E"
    if v >= 70:
        return "#D8542F"
    return "#C77A0B"


def reco_color(v) -> str:
    s = str(v).lower()
    if "strong buy" in s or s == "buy":
        return "#0B7A4B"
    if "sell" in s:
        return "#B3261E"
    if "hold" in s:
        return "#C77A0B"
    return "#8794A1"


def reco_info(v) -> str:
    return RECO_INFO.get(str(v), "")


# --------------------------- health scorecard -----------------------------

def _find_row(fin, *keys):
    if fin is None or getattr(fin, "empty", True):
        return None
    for k in keys:
        for idx in fin.index:
            if k.lower() in str(idx).lower():
                return fin.loc[idx]
    return None


def _vals(series, n=1):
    if series is None:
        return None
    s = pd.Series(series).dropna()
    if s.empty:
        return None
    return s.sort_index(ascending=False).head(n).astype(float)


def _band(v, good, ok, higher_better=True):
    if v is None:
        return "No Data"
    if higher_better:
        return "Yes" if v >= good else ("Neutral" if v >= ok else "No")
    return "Yes" if v <= good else ("Neutral" if v <= ok else "No")


def health_checks(income, balance, cash) -> list[dict]:
    out = []

    def add(name, verdict, detail=""):
        out.append({"name": name, "verdict": verdict, "detail": detail})

    ni = _find_row(income, "Net Income")
    cfo = _find_row(cash, "Operating Cash Flow", "Total Cash From Operating")
    n3, c3 = _vals(ni, 3), _vals(cfo, 3)
    if n3 is not None and c3 is not None and n3.sum() != 0:
        r = c3.sum() / n3.sum()
        add("Cash flow backs reported profit", _band(r, 0.8, 0.5), f"{r:.2f}x")
    else:
        add("Cash flow backs reported profit", "No Data")

    c1 = _vals(cfo, 1)
    add("Operating cash flow positive",
        "Yes" if c1 is not None and c1.iloc[0] > 0 else ("No" if c1 is not None else "No Data"),
        f"Rs {c1.iloc[0]/1e7:,.0f} Cr" if c1 is not None else "")

    cx = _vals(_find_row(cash, "Capital Expenditure"), 1)
    if c1 is not None and cx is not None:
        fcf = c1.iloc[0] - abs(cx.iloc[0])
        add("Free cash flow positive", "Yes" if fcf > 0 else "No", f"Rs {fcf/1e7:,.0f} Cr")
    else:
        add("Free cash flow positive", "No Data")

    td = _vals(_find_row(balance, "Total Debt"), 2)
    eq = _vals(_find_row(balance, "Stockholders Equity", "Total Equity"), 1)
    if td is not None and eq is not None and eq.iloc[0]:
        de = td.iloc[0] / eq.iloc[0]
        add("Debt to equity", _band(de, 0.5, 1.5, False), f"{de:.2f}x")
    else:
        add("Debt to equity", "No Data")

    if td is not None and len(td) == 2:
        add("Debt reduced vs last year", "Yes" if td.iloc[0] < td.iloc[1] else "No",
            f"{td.iloc[0]/1e7:,.0f} Cr")
    else:
        add("Debt reduced vs last year", "No Data")

    ebit = _vals(_find_row(income, "EBIT", "Operating Income"), 1)
    ie = _vals(_find_row(income, "Interest Expense"), 1)
    if ebit is not None and ie is not None and ie.iloc[0]:
        cov = ebit.iloc[0] / abs(ie.iloc[0])
        add("Interest coverage", _band(cov, 5, 2), f"{cov:.1f}x")
    else:
        add("Interest coverage", "No Data")

    ca = _vals(_find_row(balance, "Current Assets"), 1)
    cl = _vals(_find_row(balance, "Current Liabilities"), 1)
    if ca is not None and cl is not None and cl.iloc[0]:
        cr = ca.iloc[0] / cl.iloc[0]
        add("Current ratio", _band(cr, 1.5, 1.0), f"{cr:.2f}x")
    else:
        add("Current ratio", "No Data")

    rev = _vals(_find_row(income, "Total Revenue"), 3)
    rg = (rev.iloc[0] / rev.iloc[1] - 1) if rev is not None and len(rev) >= 2 and rev.iloc[1] else None
    for label, ser in (("Receivables in line with sales",
                        _vals(_find_row(balance, "Receivables", "Accounts Receivable"), 2)),
                       ("Inventory in line with sales",
                        _vals(_find_row(balance, "Inventory"), 2))):
        if ser is not None and len(ser) == 2 and ser.iloc[1] and rg is not None:
            g = ser.iloc[0] / ser.iloc[1] - 1
            add(label, _band(g - rg, 0.10, 0.25, False), f"{g*100:.0f}% vs {rg*100:.0f}%")
        else:
            add(label, "No Data")

    if rev is not None and len(rev) >= 3 and rev.iloc[2]:
        add("Revenue growing (3y)", "Yes" if rev.iloc[0] > rev.iloc[2] else "No",
            f"{(rev.iloc[0]/rev.iloc[2]-1)*100:.0f}%")
    else:
        add("Revenue growing (3y)", "No Data")

    if n3 is not None and len(n3) >= 3 and n3.iloc[2]:
        add("Profit growing (3y)", "Yes" if n3.iloc[0] > n3.iloc[2] else "No",
            f"{(n3.iloc[0]/n3.iloc[2]-1)*100:.0f}%")
        add("Profitable every year (3y)", "Yes" if (n3 > 0).all() else "No")
    else:
        add("Profit growing (3y)", "No Data")
        add("Profitable every year (3y)", "No Data")

    if n3 is not None and eq is not None and eq.iloc[0]:
        roe = n3.iloc[0] / eq.iloc[0] * 100
        add("Return on equity", _band(roe, 15, 8), f"{roe:.1f}%")
    else:
        add("Return on equity", "No Data")

    om = _vals(_find_row(income, "Operating Income", "EBIT"), 2)
    if (om is not None and rev is not None and len(om) == 2 and len(rev) >= 2
            and rev.iloc[0] and rev.iloc[1]):
        m0, m1 = om.iloc[0] / rev.iloc[0], om.iloc[1] / rev.iloc[1]
        add("Operating margin holding up", "Yes" if m0 >= m1 else "No",
            f"{m0*100:.1f}% vs {m1*100:.1f}%")
    else:
        add("Operating margin holding up", "No Data")

    if n3 is not None and rev is not None and rev.iloc[0]:
        nm = n3.iloc[0] / rev.iloc[0] * 100
        add("Net margin positive", "Yes" if nm > 0 else "No", f"{nm:.1f}%")
    else:
        add("Net margin positive", "No Data")

    return out


# ------------------------------- universe ---------------------------------

def _fetch_symbol_csv(url: str) -> list[str]:
    hdr = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Accept": "text/csv,*/*"}
    try:
        r = requests.get(url, headers=hdr, timeout=20)
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))
        col = next(c for c in df.columns if c.strip().upper() == "SYMBOL")
        return [f"{s.strip()}.NS" for s in df[col].dropna().astype(str) if s.strip()]
    except Exception:
        return []


@st.cache_data(ttl=86400, show_spinner=False)
def load_all_nse_tickers() -> list[str]:
    return _fetch_symbol_csv(_EQUITY_LIST_URL)


@st.cache_data(ttl=86400, show_spinner=False)
def load_nifty500_tickers() -> list[str]:
    return _fetch_symbol_csv(_INDEX500_LIST_URL)


@st.cache_data(ttl=86400, show_spinner=False)
def load_face_values() -> dict:
    """Map SYMBOL -> face value from NSE's public EQUITY_L.csv (no auth needed)."""
    hdr = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Accept": "text/csv,*/*"}
    try:
        r = requests.get(_EQUITY_LIST_URL, headers=hdr, timeout=20)
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))
        df.columns = [c.strip().upper() for c in df.columns]
        sym_c = next((c for c in df.columns if c == "SYMBOL"), None)
        fv_c = next((c for c in df.columns if c == "FACE VALUE"), None)
        if not sym_c or not fv_c:
            return {}
        out = {}
        for _, row in df[[sym_c, fv_c]].dropna().iterrows():
            try:
                out[str(row[sym_c]).strip().upper()] = float(row[fv_c])
            except Exception:
                continue
        return out
    except Exception:
        return {}


def parse_uploaded_symbols(text: str) -> list[str]:
    out = []
    for line in text.replace(",", "\n").splitlines():
        s = line.strip().upper()
        if not s or s == "SYMBOL":
            continue
        out.append(s if s.endswith(".NS") else f"{s}.NS")
    return out


def tickers_for(universe_choice: str, uploaded: str | None = None) -> list[str]:
    if universe_choice == "NIFTY 50":
        return list(NIFTY50)
    if universe_choice == "NIFTY 500":
        return load_nifty500_tickers()
    if universe_choice == "All NIFTY Stocks":
        return load_all_nse_tickers()
    return []


SCAN_SEP = "|"


def _cfg_str(u: str, th, n) -> str:
    return f"{u}{SCAN_SEP}{th}{SCAN_SEP}{n}"


# ----------------------------- data fetching ------------------------------

def fetch_fundamentals(symbol: str) -> dict:
    name = symbol.replace(".NS", "")
    sector, reco, target, pe = "n/a", "n/a", "n/a", None
    tk = yf.Ticker(symbol)
    try:
        info = tk.info or {}
        name = info.get("shortName") or info.get("longName") or name
        if info.get("sector"):
            sector = str(info["sector"]).strip()
        rk = info.get("recommendationKey")
        if rk and rk.lower() != "none":
            reco = rk.replace("_", " ").title()
        tp = info.get("targetMeanPrice")
        if tp:
            target = round(float(tp), 2)
        pev = info.get("trailingPE")
        if pev:
            pe = round(float(pev), 2)
    except Exception:
        pass
    if target == "n/a":
        try:
            m = (tk.analyst_price_targets or {}).get("mean")
            if m:
                target = round(float(m), 2)
        except Exception:
            pass
    return {"name": name, "sector": sector, "reco": reco, "target": target, "pe": pe}


@st.cache_data(ttl=60, show_spinner=False)
def fetch_detail(symbol: str) -> dict:
    out = {"name": symbol.replace(".NS", ""), "sector": "n/a", "mcap": None,
           "price": None, "high52": None, "low52": None, "rsi": None,
           "reco": "n/a", "target": "n/a", "levels": {}, "swing": (None, None),
           "checks": [], "face": None, "book": None, "beta": None, "error": None}
    try:
        tk = yf.Ticker(symbol)
        hist = tk.history(period="2y", interval="1d", auto_adjust=False)
        if hist is None or hist.empty:
            out["error"] = "No price history found for this symbol. Check the spelling."
            return out
        close = hist["Close"].dropna()
        out["price"] = float(close.iloc[-1])
        out["rsi"] = compute_rsi(close)
        yr = hist.tail(252)
        out["high52"], out["low52"] = float(yr["High"].max()), float(yr["Low"].min())

        monthly = hist.resample("ME").agg({"High": "max", "Low": "min", "Close": "last"}).dropna()
        if len(monthly) >= 2:
            prev = monthly.iloc[-2]
            out["levels"] = pivot_levels(float(prev["High"]), float(prev["Low"]),
                                         float(prev["Close"]))
        out["swing"] = swing_levels(close.tail(180), out["price"])

        try:
            info = tk.info or {}
            out["name"] = info.get("shortName") or info.get("longName") or out["name"]
            out["mcap"] = info.get("marketCap")
            if info.get("sector"):
                out["sector"] = str(info["sector"]).strip()
            rk = info.get("recommendationKey")
            if rk and rk.lower() != "none":
                out["reco"] = rk.replace("_", " ").title()
            tp = info.get("targetMeanPrice")
            if tp:
                out["target"] = round(float(tp), 2)
            out["book"] = info.get("bookValue")
            out["beta"] = info.get("beta")
            out["face"] = info.get("faceValue") or info.get("parValue")
            if out["face"] is None:
                out["face"] = load_face_values().get(symbol.replace(".NS", "").upper())
        except Exception:
            pass
        try:
            out["checks"] = health_checks(tk.income_stmt, tk.balance_sheet, tk.cashflow)
        except Exception:
            out["checks"] = []
    except Exception as exc:
        out["error"] = f"Could not load this stock ({type(exc).__name__}). Try again shortly."
    return out


@st.cache_data(ttl=900, show_spinner=False)
def fetch_news(limit: int = 8) -> list[dict]:
    """Latest market headlines from public RSS feeds (headline + link only)."""
    feeds = [
        ("Economic Times", "https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms"),
        ("Moneycontrol", "https://www.moneycontrol.com/rss/marketreports.xml"),
        ("Business Standard", "https://www.business-standard.com/rss/markets-106.rss"),
    ]
    out = []
    for source, url in feeds:
        try:
            r = requests.get(url, timeout=6, headers={"User-Agent": "Mozilla/5.0"})
            root = ET.fromstring(r.content)
            for item in root.iter("item"):
                title = (item.findtext("title") or "").strip()
                link = (item.findtext("link") or "").strip()
                if title and link:
                    out.append({"title": title, "url": link, "source": source})
        except Exception:
            continue
    return out[:limit]


IST = dt.timezone(dt.timedelta(hours=5, minutes=30))


def market_open() -> bool:
    now = dt.datetime.now(IST)
    if now.weekday() >= 5:
        return False
    return dt.time(9, 15) <= now.time() <= dt.time(15, 30)


@st.cache_data(ttl=20, show_spinner=False)
def fetch_indices(bucket: int = 0) -> list[dict]:
    specs = [("NIFTY 50", "^NSEI"), ("SENSEX", "^BSESN"),
             ("BANK NIFTY", "^NSEBANK"), ("NIFTY IT", "^CNXIT"),
             ("NIFTY MIDCAP 100", "^NSEMDCP50")]
    out = []
    for name, tk in specs:
        try:
            h = yf.Ticker(tk).history(period="5d")["Close"].dropna()
            if len(h) >= 2:
                val = float(h.iloc[-1])
                chg = (val / float(h.iloc[-2]) - 1) * 100
                out.append({"name": name, "value": val, "chg": chg})
        except Exception:
            continue
    return out


@st.cache_data(ttl=300, show_spinner=False)
def scan(tickers: tuple[str, ...], threshold: float, fetch_fund_limit: int) -> pd.DataFrame:
    bar = st.progress(0.0, text="Downloading price history...")
    close_map, high_map, vol_map, volser_map = {}, {}, {}, {}
    batch, n = 200, len(tickers)
    for i in range(0, n, batch):
        chunk = list(tickers[i:i + batch])
        data = yf.download(chunk, period="1y", interval="1d", group_by="ticker",
                           auto_adjust=False, threads=True, progress=False)
        for sym in chunk:
            try:
                df = data[sym] if len(chunk) > 1 else data
                c = df["Close"].dropna()
                if c.empty:
                    continue
                close_map[sym] = c
                high_map[sym] = float(df["High"].dropna().max())
                v = df["Volume"].dropna() if "Volume" in df else None
                vol_map[sym] = float(v.iloc[-1]) if v is not None and not v.empty else None
                volser_map[sym] = v
            except Exception:
                continue
        bar.progress(min((i + batch) / n, 1.0), text=f"Downloaded {min(i+batch, n)}/{n} stocks")

    hits = []
    for sym, c in close_map.items():
        rsi = compute_rsi(c)
        if rsi is not None and rsi >= threshold:
            hits.append((sym, rsi, float(c.iloc[-1]), high_map.get(sym)))
    hits.sort(key=lambda x: x[1], reverse=True)

    def _sma(series, n):
        if series is None or len(series) < n:
            return None
        return round(float(series.rolling(n).mean().iloc[-1]), 2)

    rows, total = [], min(len(hits), fetch_fund_limit)
    bar.progress(0.0, text=f"Loading details for {total} matches")
    for idx, (sym, rsi, price, high52) in enumerate(hits):
        if idx < fetch_fund_limit:
            f = fetch_fundamentals(sym)
            time.sleep(0.4)
            bar.progress((idx + 1) / max(total, 1), text=f"Loaded {idx+1}/{total}")
        else:
            f = {"name": sym.replace(".NS", ""), "sector": "not loaded",
                 "reco": "not loaded", "target": "not loaded", "pe": None}
        cser = close_map.get(sym)
        _vw = compute_vwap(cser, volser_map.get(sym))
        _vwap_pos = ("Above" if price > _vw else "Below") if _vw else "n/a"
        rows.append({
            "Date": dt.date.today().strftime("%d-%m-%Y"),
            "Stock Symbol": sym.replace(".NS", ""),
            "Sector": f["sector"],
            "Current Price (Rs)": round(price, 2),
            "Volume": vol_map.get(sym),
            "52 Week High (Rs)": round(high52, 2) if high52 else None,
            "RSI": round(rsi, 1),
            "VWAP": _vwap_pos,
            "PE": f.get("pe"),
            "Sec PE": None,
            "Buy/Sell": f["reco"],
            "1 Year Target (Rs)": f["target"],
            "_SMA20": _sma(cser, 20),
            "_SMA50": _sma(cser, 50),
            "_SMA200": _sma(cser, 200),
        })
    # Sector P/E = mean trailing P/E of scanned stocks in the same sector
    sec_pes: dict[str, list[float]] = {}
    for r in rows:
        if isinstance(r["PE"], (int, float)) and isinstance(r["Sector"], str):
            sec_pes.setdefault(r["Sector"], []).append(float(r["PE"]))
    sec_avg = {k: round(sum(v) / len(v), 2) for k, v in sec_pes.items()}
    for r in rows:
        r["Sec PE"] = sec_avg.get(r["Sector"])
    bar.empty()
    return pd.DataFrame(rows, columns=HEADERS + ["_SMA20", "_SMA50", "_SMA200"])


# =========================== screening engine =============================
# Architecture (two-stage funnel):
#   stage 1  bulk_ohlcv()  one batched 1y OHLCV download for the whole universe
#            -> tech_row() computes every technical value locally (free, fast)
#   stage 2  fund_row()    one .info call per SURVIVING symbol, capped, cached
#            6h -- the expensive half only runs on stocks that already passed
#            the technical filters.
# Both stages feed the same flat row dict, so one filter engine
# (apply_num_filters) serves the custom screener, and tech_row alone serves
# the ETF tab. Mutual funds have no OHLCV, so they use their own AMFI +
# mfapi.in pair with the same table/export helpers.

CUSTOM_FUND_CAP = 120

# (column, label, group T=technical F=fundamental, decimals)
NUM_COLS = [
    ("Price", "Price (Rs)", "T", 2),
    ("Chg1D", "1 day %", "T", 2),
    ("Chg1W", "1 week %", "T", 2),
    ("Chg1M", "1 month %", "T", 2),
    ("Chg3M", "3 month %", "T", 2),
    ("Chg6M", "6 month %", "T", 2),
    ("Chg1Y", "1 year %", "T", 2),
    ("RSI", "RSI (14)", "T", 1),
    ("SMA20", "20 DMA", "T", 2),
    ("SMA50", "50 DMA", "T", 2),
    ("SMA200", "200 DMA", "T", 2),
    ("Vs200", "Price vs 200 DMA %", "T", 2),
    ("VWAP20", "VWAP (20d)", "T", 2),
    ("VsVWAP", "Price vs VWAP %", "T", 2),
    ("High52", "52 week high", "T", 2),
    ("Low52", "52 week low", "T", 2),
    ("FromHigh", "Below 52w high %", "T", 2),
    ("FromLow", "Above 52w low %", "T", 2),
    ("Volume", "Volume", "T", 0),
    ("AvgVol20", "Avg volume (20d)", "T", 0),
    ("VolX", "Volume vs avg (x)", "T", 2),
    ("ATRpct", "ATR 14 %", "T", 2),
    ("MCapCr", "Market cap (Rs cr)", "F", 0),
    ("PE", "P/E", "F", 2),
    ("FwdPE", "Forward P/E", "F", 2),
    ("PB", "P/B", "F", 2),
    ("ROE", "ROE %", "F", 2),
    ("ROA", "ROA %", "F", 2),
    ("DE", "Debt / equity", "F", 2),
    ("DivYld", "Dividend yield %", "F", 2),
    ("EPS", "EPS (Rs)", "F", 2),
    ("RevGrowth", "Revenue growth %", "F", 2),
    ("ProfitGrowth", "Earnings growth %", "F", 2),
    ("OpMargin", "Operating margin %", "F", 2),
    ("NetMargin", "Net margin %", "F", 2),
    ("CurrRatio", "Current ratio", "F", 2),
    ("Beta", "Beta", "F", 2),
    ("Target", "1Y target (Rs)", "F", 2),
    ("Upside", "Upside to target %", "F", 2),
]
COL_META = {c: (lab, grp, dp) for c, lab, grp, dp in NUM_COLS}
COL_META["iNAV"] = ("iNAV (Rs)", "T", 2)      # ETF-only, not a stock screen filter
COL_META["PremDisc"] = ("Prem / disc %", "T", 2)
FUND_KEYS = {c for c, _l, g, _d in NUM_COLS if g == "F"}

PRESETS = {
    "Oversold quality": [("RSI", "at most", 35, None), ("ROE", "at least", 12, None),
                         ("DE", "at most", 1.0, None)],
    "Momentum breakout": [("RSI", "at least", 60, None), ("Vs200", "at least", 5, None),
                          ("FromHigh", "at most", 8, None), ("VolX", "at least", 1.2, None)],
    "Value screen": [("PE", "between", 0, 18), ("PB", "at most", 3,  None),
                     ("DivYld", "at least", 1, None)],
    "Quality compounder": [("ROE", "at least", 18, None), ("DE", "at most", 0.5, None),
                           ("ProfitGrowth", "at least", 10, None),
                           ("OpMargin", "at least", 15, None)],
    "Beaten down": [("FromHigh", "at least", 30, None), ("RSI", "at most", 45, None)],
}


@st.cache_data(ttl=900, show_spinner=False)
def bulk_ohlcv(tickers: tuple[str, ...], bucket: int = 0) -> dict:
    """One batched 1-year daily OHLCV download for a whole universe. `bucket` is
    a time slot -- pass int(time.time() // 60) to force a fresh pull each minute
    while the market is open."""
    out: dict = {}
    batch, n = 200, len(tickers)
    bar = st.progress(0.0, text="Downloading price history...")
    for i in range(0, n, batch):
        chunk = list(tickers[i:i + batch])
        try:
            data = yf.download(chunk, period="1y", interval="1d", group_by="ticker",
                               auto_adjust=False, threads=True, progress=False)
        except Exception:
            data = None
        if data is not None and not getattr(data, "empty", True):
            for sym in chunk:
                try:
                    df = data[sym] if len(chunk) > 1 else data
                    c = df["Close"].dropna()
                    if len(c) < 30:
                        continue
                    out[sym] = {
                        "close": c,
                        "high": df["High"].dropna() if "High" in df else None,
                        "low": df["Low"].dropna() if "Low" in df else None,
                        "vol": df["Volume"].dropna() if "Volume" in df else None,
                    }
                except Exception:
                    continue
        bar.progress(min((i + batch) / n, 1.0),
                     text=f"Priced {min(i + batch, n)}/{n} securities")
    bar.empty()
    return out


def _numv(v, dp=2):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f != f or f in (float("inf"), float("-inf")):
        return None
    return round(f, dp)


def _chg_pct(c: pd.Series, sessions: int):
    if c is None or len(c) <= sessions:
        return None
    base = float(c.iloc[-1 - sessions])
    return _numv((float(c.iloc[-1]) / base - 1) * 100) if base else None


def _atr_pct(h, l, c, n: int = 14):
    if h is None or l is None or c is None or len(c) < n + 1:
        return None
    try:
        df = pd.DataFrame({"h": h, "l": l, "c": c}).dropna()
        pc = df["c"].shift(1)
        tr = pd.concat([df["h"] - df["l"], (df["h"] - pc).abs(),
                        (df["l"] - pc).abs()], axis=1).max(axis=1)
        atr = float(tr.rolling(n).mean().iloc[-1])
        last = float(df["c"].iloc[-1])
        return _numv(atr / last * 100) if last else None
    except Exception:
        return None


def tech_row(symbol: str, d: dict) -> dict:
    """Every technical value we can derive from one OHLCV history."""
    c, v, h, l = d.get("close"), d.get("vol"), d.get("high"), d.get("low")
    price = float(c.iloc[-1])
    yr = c.tail(252)
    hi = float(h.tail(252).max()) if h is not None and not h.empty else float(yr.max())
    lo = float(l.tail(252).min()) if l is not None and not l.empty else float(yr.min())

    def sma(n):
        return _numv(c.rolling(n).mean().iloc[-1]) if len(c) >= n else None

    s200, vw = sma(200), compute_vwap(c, v)
    avg20 = _numv(v.rolling(20).mean().iloc[-1], 0) if v is not None and len(v) >= 20 else None
    last_vol = _numv(v.iloc[-1], 0) if v is not None and not v.empty else None
    return {
        "Symbol": symbol.replace(".NS", ""),
        "Price": _numv(price), "RSI": _numv(compute_rsi(c), 1),
        "Chg1D": _chg_pct(c, 1), "Chg1W": _chg_pct(c, 5), "Chg1M": _chg_pct(c, 21),
        "Chg3M": _chg_pct(c, 63), "Chg6M": _chg_pct(c, 126), "Chg1Y": _chg_pct(c, 248),
        "SMA20": sma(20), "SMA50": sma(50), "SMA200": s200,
        "Vs200": _numv((price / s200 - 1) * 100) if s200 else None,
        "VWAP20": _numv(vw), "VsVWAP": _numv((price / vw - 1) * 100) if vw else None,
        "High52": _numv(hi), "Low52": _numv(lo),
        "FromHigh": _numv((1 - price / hi) * 100) if hi else None,
        "FromLow": _numv((price / lo - 1) * 100) if lo else None,
        "Volume": last_vol, "AvgVol20": avg20,
        "VolX": _numv(last_vol / avg20) if (last_vol and avg20) else None,
        "ATRpct": _atr_pct(h, l, c),
    }


@st.cache_data(ttl=21600, show_spinner=False)
def fund_row(symbol: str) -> dict:
    """Fundamentals for one symbol. Every value is reported as published --
    missing figures stay None and are shown as 'n/a', never guessed."""
    out = {k: None for k in FUND_KEYS}
    out.update({"Sector": "n/a", "Industry": "n/a", "Buy/Sell": "n/a"})
    try:
        info = yf.Ticker(symbol).info or {}
    except Exception:
        return out
    if not info:
        return out

    def pct(key):
        v = info.get(key)
        return _numv(float(v) * 100) if isinstance(v, (int, float)) else None

    mcap = info.get("marketCap")
    out["MCapCr"] = _numv(float(mcap) / 1e7, 0) if mcap else None
    out["PE"], out["FwdPE"] = _numv(info.get("trailingPE")), _numv(info.get("forwardPE"))
    out["PB"] = _numv(info.get("priceToBook"))
    out["ROE"], out["ROA"] = pct("returnOnEquity"), pct("returnOnAssets")
    de = info.get("debtToEquity")
    out["DE"] = _numv(float(de) / 100) if isinstance(de, (int, float)) else None
    dy = info.get("dividendYield")
    if isinstance(dy, (int, float)):
        out["DivYld"] = _numv(float(dy) * 100) if float(dy) < 1 else _numv(dy)
    out["EPS"] = _numv(info.get("trailingEps"))
    out["RevGrowth"] = pct("revenueGrowth")
    out["ProfitGrowth"] = pct("earningsGrowth") or pct("earningsQuarterlyGrowth")
    out["OpMargin"], out["NetMargin"] = pct("operatingMargins"), pct("profitMargins")
    out["CurrRatio"], out["Beta"] = _numv(info.get("currentRatio")), _numv(info.get("beta"))
    tp, price = info.get("targetMeanPrice"), info.get("currentPrice")
    out["Target"] = _numv(tp)
    if tp and price:
        out["Upside"] = _numv((float(tp) / float(price) - 1) * 100)
    if info.get("sector"):
        out["Sector"] = str(info["sector"]).strip()
    if info.get("industry"):
        out["Industry"] = str(info["industry"]).strip()
    rk = info.get("recommendationKey")
    if rk and str(rk).lower() != "none":
        out["Buy/Sell"] = str(rk).replace("_", " ").title()
    return out


def apply_num_filters(df: pd.DataFrame, filters: list[dict]) -> pd.DataFrame:
    for f in filters:
        col = f["col"]
        if col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce")
        if f["op"] == "at least":
            df = df[s >= f["v1"]]
        elif f["op"] == "at most":
            df = df[s <= f["v1"]]
        else:
            lo, hi = sorted([f["v1"], f["v2"]])
            df = df[(s >= lo) & (s <= hi)]
    return df


def custom_screen(tickers: tuple[str, ...], filters: list[dict],
                  sectors: list[str], cap: int = CUSTOM_FUND_CAP) -> tuple[pd.DataFrame, int, bool]:
    """Run the funnel. Returns (dataframe, universe size priced, fundamentals loaded)."""
    data = bulk_ohlcv(tuple(tickers))
    if not data:
        return pd.DataFrame(), 0, False
    df = pd.DataFrame([tech_row(s, d) for s, d in data.items()])
    priced = len(df)
    df = apply_num_filters(df, [f for f in filters if COL_META[f["col"]][1] == "T"])
    need_fund = bool(sectors) or any(COL_META[f["col"]][1] == "F" for f in filters)
    if not need_fund or df.empty:
        return df.reset_index(drop=True), priced, False

    df = df.head(cap).copy()
    bar = st.progress(0.0, text=f"Loading fundamentals for {len(df)} matches")
    frows = []
    for i, sym in enumerate(df["Symbol"].tolist()):
        frows.append(fund_row(f"{sym}.NS"))
        time.sleep(0.25)
        bar.progress((i + 1) / len(df), text=f"Fundamentals {i + 1}/{len(df)}")
    bar.empty()
    fdf = pd.DataFrame(frows, index=df.index)
    df = pd.concat([df, fdf], axis=1)
    if sectors:
        df = df[df["Sector"].astype(str).isin(sectors)]
    df = apply_num_filters(df, [f for f in filters if COL_META[f["col"]][1] == "F"])
    return df.reset_index(drop=True), priced, True


# ------------------------------- mutual funds ------------------------------

_AMFI_NAV_URL = "https://www.amfiindia.com/spages/NAVAll.txt"


def _mf_type(category: str) -> str:
    c = (category or "").lower()
    for key, label in (("equity", "Equity"), ("debt", "Debt"), ("hybrid", "Hybrid"),
                       ("solution", "Solution oriented"), ("index", "Index / ETF"),
                       ("other", "Other")):
        if key in c:
            return label
    return "Other"


@st.cache_data(ttl=21600, show_spinner=False)
def load_mutual_funds() -> pd.DataFrame:
    """Every scheme AMFI publishes, with its latest declared NAV. Public file,
    no auth, refreshed once every business evening."""
    try:
        r = requests.get(_AMFI_NAV_URL, timeout=25,
                         headers={"User-Agent": "Mozilla/5.0"})
        text = r.text
    except Exception:
        return pd.DataFrame()
    rows, amc, cat = [], "", ""
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if ";" not in line:
            if "Schemes(" in line or "Scheme(" in line:
                inner = line[line.find("(") + 1:line.rfind(")")] if "(" in line else line
                cat = inner.strip() or line
            else:
                amc = line.replace("Mutual Fund", "").strip() or line
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 6 or not parts[0].isdigit():
            continue
        try:
            nav = float(parts[4])
        except (ValueError, IndexError):
            continue
        rows.append({"Code": parts[0], "Scheme": parts[3], "AMC": amc,
                     "Category": cat, "Type": _mf_type(cat), "NAV": round(nav, 4),
                     "Date": parts[5]})
    return pd.DataFrame(rows)


@st.cache_data(ttl=21600, show_spinner=False)
def mf_returns(code: str) -> dict:
    """Point-to-point returns from the scheme's own published NAV history
    (mfapi.in mirrors AMFI). >1y figures are annualised (CAGR)."""
    out = {k: None for k in ("1M", "3M", "6M", "1Y", "3Y", "5Y")}
    try:
        j = requests.get(f"https://api.mfapi.in/mf/{code}", timeout=10).json()
        data = j.get("data") or []
        pairs = {}
        for x in data:
            try:
                pairs[pd.to_datetime(x["date"], format="%d-%m-%Y")] = float(x["nav"])
            except (ValueError, TypeError, KeyError):
                continue
        if len(pairs) < 2:
            return out
        ser = pd.Series(pairs).sort_index()
    except Exception:
        return out
    last, last_dt = float(ser.iloc[-1]), ser.index[-1]
    for label, days in (("1M", 30), ("3M", 91), ("6M", 182), ("1Y", 365),
                        ("3Y", 1095), ("5Y", 1825)):
        past = ser[ser.index <= last_dt - pd.Timedelta(days=days)]
        if past.empty:
            continue
        base = float(past.iloc[-1])
        if base <= 0:
            continue
        yrs = days / 365.0
        out[label] = (_numv((last / base - 1) * 100) if yrs <= 1.05
                      else _numv(((last / base) ** (1 / yrs) - 1) * 100))
    return out


# ----------------------------------- etfs ----------------------------------
# Symbol, display name, category. Symbols that return no history are dropped
# at render time, so the list can carry newer launches safely.
ETF_LIST = [
    ("NIFTYBEES", "Nippon Nifty 50", "Broad market"),
    ("SETFNIF50", "SBI Nifty 50", "Broad market"),
    ("UTINIFTETF", "UTI Nifty 50", "Broad market"),
    ("NIFTYIETF", "ICICI Nifty 50", "Broad market"),
    ("HDFCNIFTY", "HDFC Nifty 50", "Broad market"),
    ("JUNIORBEES", "Nippon Nifty Next 50", "Broad market"),
    ("ICICIB22", "BHARAT 22", "Broad market"),
    ("EQUAL50ADD", "Nifty 50 Equal Weight", "Broad market"),
    ("MID150BEES", "Nippon Nifty Midcap 150", "Mid & small cap"),
    ("MIDCAPETF", "ICICI Nifty Midcap 150", "Mid & small cap"),
    ("MOM100", "Motilal Midcap 100", "Mid & small cap"),
    ("HDFCSML250", "HDFC Nifty Smallcap 250", "Mid & small cap"),
    ("MOSMALL250", "Motilal Smallcap 250", "Mid & small cap"),
    ("BANKBEES", "Nippon Nifty Bank", "Sector & theme"),
    ("SETFNIFBK", "SBI Nifty Bank", "Sector & theme"),
    ("BANKIETF", "ICICI Nifty Bank", "Sector & theme"),
    ("PSUBNKBEES", "Nippon PSU Bank", "Sector & theme"),
    ("ITBEES", "Nippon Nifty IT", "Sector & theme"),
    ("ITIETF", "ICICI Nifty IT", "Sector & theme"),
    ("PHARMABEES", "Nippon Nifty Pharma", "Sector & theme"),
    ("AUTOBEES", "Nippon Nifty Auto", "Sector & theme"),
    ("CONSUMBEES", "Nippon Nifty India Consumption", "Sector & theme"),
    ("INFRABEES", "Nippon Nifty Infrastructure", "Sector & theme"),
    ("FMCGIETF", "ICICI Nifty FMCG", "Sector & theme"),
    ("METALIETF", "ICICI Nifty Metal", "Sector & theme"),
    ("HEALTHIETF", "ICICI Nifty Healthcare", "Sector & theme"),
    ("PVTBANIETF", "ICICI Nifty Private Bank", "Sector & theme"),
    ("MAKEINDIA", "Nippon Nifty India Manufacturing", "Sector & theme"),
    ("MOMOMENTUM", "Motilal Nifty 200 Momentum 30", "Factor"),
    ("ALPHA", "Nippon Nifty Alpha 50", "Factor"),
    ("ALPL30IETF", "ICICI Alpha Low Vol 30", "Factor"),
    ("MOVALUE", "Motilal Nifty 500 Value 50", "Factor"),
    ("LOWVOLIETF", "ICICI Nifty Low Vol 30", "Factor"),
    ("MOQUALITY", "Motilal Nifty 200 Quality 30", "Factor"),
    ("GOLDBEES", "Nippon Gold", "Gold & silver"),
    ("SETFGOLD", "SBI Gold", "Gold & silver"),
    ("HDFCGOLD", "HDFC Gold", "Gold & silver"),
    ("AXISGOLD", "Axis Gold", "Gold & silver"),
    ("GOLDIETF", "ICICI Gold", "Gold & silver"),
    ("SILVERBEES", "Nippon Silver", "Gold & silver"),
    ("SILVERIETF", "ICICI Silver", "Gold & silver"),
    ("HDFCSILVER", "HDFC Silver", "Gold & silver"),
    ("MON100", "Motilal Nasdaq 100", "International"),
    ("MAFANG", "Mirae NYSE FANG+", "International"),
    ("MASPTOP50", "Mirae S&P 500 Top 50", "International"),
    ("HNGSNGBEES", "Nippon Hang Seng", "International"),
    ("LIQUIDBEES", "Nippon Liquid", "Debt & liquid"),
    ("GILT5YBEES", "Nippon Nifty 5Y G-Sec", "Debt & liquid"),
    ("LTGILTBEES", "Nippon Long Term Gilt", "Debt & liquid"),
    ("EBBETF0433", "Bharat Bond 2033", "Debt & liquid"),
]
ETF_META = {s: (n, c) for s, n, c in ETF_LIST}
ETF_CATEGORIES = list(dict.fromkeys(c for _s, _n, c in ETF_LIST))


@st.cache_data(ttl=60, show_spinner=False)
def fetch_nse_etf_live(bucket: int = 0) -> dict:
    """Live ETF board from NSE -- iNAV, last traded price and day change for the
    whole list in ONE request, so iNAV can refresh on the same 1-minute cycle as
    prices instead of a slow per-fund lookup."""
    hdr = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
           "Accept": "application/json, text/plain, */*",
           "Accept-Language": "en-US,en;q=0.9",
           "Referer": "https://www.nseindia.com/market-data/exchange-traded-funds-etf"}
    try:
        sess = requests.Session()
        sess.headers.update(hdr)
        sess.get("https://www.nseindia.com/market-data/exchange-traded-funds-etf",
                 timeout=8)
        rows = (sess.get("https://www.nseindia.com/api/etf", timeout=10).json()
                or {}).get("data") or []
    except Exception:
        return {}
    out: dict = {}
    for d in rows:
        sym = str(d.get("symbol", "")).strip().upper()
        if not sym:
            continue

        def num(*keys):
            for k in keys:
                v = d.get(k)
                if isinstance(v, (int, float)):
                    return float(v)
                if isinstance(v, str):
                    try:
                        return float(v.replace(",", "").strip())
                    except ValueError:
                        continue
            return None

        out[sym] = {"inav": num("nav", "iNav", "inav"),
                    "ltp": num("ltP", "lastPrice", "ltp"),
                    "chg": num("per", "pChange"),
                    "prev": num("prevClose", "previousClose")}
    return out


@st.cache_data(ttl=300, show_spinner=False)
def etf_inav(symbols: tuple[str, ...], bucket: int = 0) -> dict:
    """Fallback iNAV, one lookup per fund, used only for ETFs the NSE board did
    not return. Missing values stay None and render as 'n/a', never the market
    price dressed up as NAV."""
    out: dict = {}
    if not symbols:
        return out
    bar = st.progress(0.0, text="Reading ETF NAVs...")
    for i, sym in enumerate(symbols):
        try:
            info = yf.Ticker(sym).info or {}
            out[sym.replace(".NS", "")] = _numv(info.get("navPrice")
                                                or info.get("netAssetValue"))
        except Exception:
            out[sym.replace(".NS", "")] = None
        bar.progress((i + 1) / len(symbols), text=f"NAVs {i + 1}/{len(symbols)}")
    bar.empty()
    return out


@st.cache_data(ttl=60, show_spinner=False)
def scan_etfs(price_bucket: int = 0, nav_bucket: int = 0) -> tuple[pd.DataFrame, bool]:
    """ETF technicals, live price and iNAV. Returns (rows, iNAV came from the
    live NSE board)."""
    data = bulk_ohlcv(tuple(f"{s}.NS" for s, _n, _c in ETF_LIST), price_bucket)
    live = fetch_nse_etf_live(price_bucket)
    missing = tuple(s for s in data
                    if not (live.get(s.replace(".NS", "").upper()) or {}).get("inav"))
    navs = etf_inav(missing, nav_bucket)
    rows = []
    for sym, d in data.items():
        r = tech_row(sym, d)
        name, cat = ETF_META.get(r["Symbol"], ("", "Other"))
        r["Name"], r["Category"] = name, cat
        lv = live.get(r["Symbol"].upper()) or {}
        if lv.get("ltp"):
            r["Price"] = _numv(lv["ltp"])
        if lv.get("chg") is not None:
            r["Chg1D"] = _numv(lv["chg"])
        r["iNAV"] = lv.get("inav") or navs.get(r["Symbol"])
        r["PremDisc"] = (_numv((r["Price"] / r["iNAV"] - 1) * 100)
                         if r["iNAV"] and r["Price"] else None)
        rows.append(r)
    df = pd.DataFrame(rows)
    if df.empty:
        return df, bool(live)
    return df.sort_values("Symbol").reset_index(drop=True), bool(live)


# ------------------------- shared table + export ---------------------------

def fmt_val(v, dp=2, pct=False):
    if v is None or (isinstance(v, float) and v != v):
        return "n/a"
    if isinstance(v, (int, float)):
        s = f"{v:,.{dp}f}"
        return f"{s}%" if pct else s
    return _html.escape(str(v))


def sh_table(headers: list[tuple[str, str]], rows: list[str]) -> str:
    head = "".join(f'<th class="th-{a}">{_html.escape(h)}</th>' for h, a in headers)
    return ('<div class="sh-tablewrap"><table class="sh-table"><thead><tr>'
            + head + "</tr></thead><tbody>" + "".join(rows) + "</tbody></table></div>")


def signed_cell(v, dp=2, suffix="%"):
    if not isinstance(v, (int, float)):
        return '<td class="c-num c-muted">n/a</td>'
    clr = "#0B7A4B" if v >= 0 else "#B3261E"
    return (f'<td class="c-num" style="color:{clr};font-weight:600">'
            f'{v:+,.{dp}f}{suffix}</td>')


def df_to_excel_bytes(df: pd.DataFrame, sheet: str, labels: dict | None = None) -> bytes:
    labels = labels or {}
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    cols = [c for c in df.columns if not c.startswith("_")]
    ws.append([labels.get(c, c) for c in cols])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="111C2B")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for _, r in df.iterrows():
        ws.append([(None if (isinstance(r[c], float) and r[c] != r[c]) else r[c])
                   for c in cols])
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Arial", size=11)
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            max(11, min(38, len(str(labels.get(c, c))) + 4))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(ws.max_row, 2)}"
    ws.cell(row=ws.max_row + 2, column=1, value="Disclaimer: " + DISCLAIMER).font = \
        Font(name="Arial", italic=True, size=9)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------- excel ----------------------------------

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RSI Screener"
    hfill = PatternFill("solid", fgColor="111C2B")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    bfont = Font(name="Arial", size=11)
    hot = PatternFill("solid", fgColor="FDEBE4")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_of = {h: i + 1 for i, h in enumerate(HEADERS)}
    for _, r in df.iterrows():
        ws.append([r[h] for h in HEADERS])
        i = ws.max_row
        for cell in ws[i]:
            cell.font = bfont
        for hh in ("Current Price (Rs)", "52 Week High (Rs)", "1 Year Target (Rs)",
                   "PE", "Sec PE"):
            c = ws.cell(row=i, column=col_of[hh])
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"
        vcell = ws.cell(row=i, column=col_of["Volume"])
        if isinstance(vcell.value, (int, float)):
            vcell.number_format = "#,##0"
        ws.cell(row=i, column=col_of["RSI"]).number_format = "0.0"
        if isinstance(r["RSI"], (int, float)) and r["RSI"] >= 70:
            for cell in ws[i]:
                cell.fill = hot

    for i, w in enumerate([12, 16, 22, 16, 14, 8, 8, 10, 14, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(ws.max_row, 2)}"
    ws.cell(row=ws.max_row + 2, column=1, value="Disclaimer: " + DISCLAIMER).font = \
        Font(name="Arial", italic=True, size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------ detail view -------------------------------

def _stat(label: str, value: str, note: str = "", tone: str = "") -> str:
    n = f'<div class="stat-n">{note}</div>' if note else ""
    return (f'<div class="stat {tone}"><div class="stat-k">{label}</div>'
            f'<div class="stat-v">{value}</div>{n}</div>')


def render_detail(symbol: str):
    d = fetch_detail(f"{symbol}.NS")
    if d["error"]:
        st.warning(d["error"])
        return

    st.markdown(f"### {_html.escape(str(symbol))}")
    st.markdown(
        f'<span style="color:{sector_color(d["sector"])};font-weight:600">'
        f'{_html.escape(str(d["sector"]))}</span> · {_html.escape(str(d["name"]))}',
        unsafe_allow_html=True)

    price = f"Rs {d['price']:,.2f}" if d["price"] else "n/a"
    mcap = f"Rs {d['mcap']/1e7:,.0f} Cr" if d["mcap"] else "n/a"
    rsi = f"{d['rsi']:.1f}" if d["rsi"] else "n/a"
    rng = f"{d['low52']:,.0f} – {d['high52']:,.0f}" if d["high52"] else "n/a"
    tgt = f"Rs {d['target']:,.2f}" if isinstance(d["target"], (int, float)) else "n/a"
    face = f"Rs {d['face']:,.2f}" if isinstance(d["face"], (int, float)) else "n/a"
    book = f"Rs {d['book']:,.2f}" if isinstance(d["book"], (int, float)) else "n/a"
    beta = f"{d['beta']:.2f}" if isinstance(d["beta"], (int, float)) else "n/a"

    cards = "".join([
        _stat("Price", price, "", ""),
        _stat("Market cap", mcap, classify_market_cap(d["mcap"]), "v"),
        _stat("RSI (14d)", rsi,
              "overbought zone" if d["rsi"] and d["rsi"] >= 70 else "momentum", "a"),
        _stat("52-week range", rng, "", "b"),
        _stat("1-year forecast", tgt,
              f"mean analyst target · {d['reco']}", "g"),
        _stat("Face value", face, "", "v"),
        _stat("Book value", book, "per share", "b"),
        _stat("Beta", beta, "volatility vs market", "a"),
    ])
    st.markdown(f'<div class="stat-row">{cards}</div>', unsafe_allow_html=True)
    st.caption(f"Live price as of {dt.datetime.now().strftime('%d-%m-%Y, %H:%M')} "
               "(server time).")

    tab_levels, tab_health = st.tabs(["Price levels", "Financial health"])

    with tab_levels:
        left, right = st.columns(2)
        with left:
            st.markdown('<div class="sec-label">Monthly pivot levels</div>',
                        unsafe_allow_html=True)
            if d["levels"]:
                lv, p = d["levels"], d["price"]
                tone = {"R3": "r", "R2": "r2", "R1": "r2", "Pivot": "p",
                        "S1": "s2", "S2": "s2", "S3": "s"}
                html, placed = [], False
                for k in ["R3", "R2", "R1", "Pivot", "S1", "S2", "S3"]:
                    if not placed and p and lv[k] < p:
                        html.append(f'<div class="lvl now"><span>Current price</span>'
                                    f'<span>{p:,.2f}</span></div>')
                        placed = True
                    html.append(f'<div class="lvl {tone[k]}"><span>{k}</span>'
                                f'<span>{lv[k]:,.2f}</span></div>')
                if not placed and p:
                    html.append(f'<div class="lvl now"><span>Current price</span>'
                                f'<span>{p:,.2f}</span></div>')
                st.markdown("".join(html), unsafe_allow_html=True)
                st.caption("Floor-trader pivots from the last completed month.")
            else:
                st.write("Not enough history to compute pivots.")
        with right:
            st.markdown('<div class="sec-label">Nearest swing levels</div>',
                        unsafe_allow_html=True)
            sup, res = d["swing"]
            st.markdown(
                f'<div class="lvl r"><span>Resistance</span><span>'
                f'{f"{res:,.2f}" if res else "none in range"}</span></div>'
                f'<div class="lvl now"><span>Current price</span><span>'
                f'{d["price"]:,.2f}</span></div>'
                f'<div class="lvl s"><span>Support</span><span>'
                f'{f"{sup:,.2f}" if sup else "none in range"}</span></div>',
                unsafe_allow_html=True)
            st.caption("Local highs and lows over roughly the last six months.")

    with tab_health:
        checks = d["checks"]
        if not checks:
            st.write("No financial statements available for this stock.")
        else:
            cnt = Counter(c["verdict"] for c in checks)
            st.markdown(
                '<div class="score-row">'
                f'<div class="score pos"><div class="score-n">{cnt.get("Yes",0)}</div>'
                '<div class="score-l">Positive</div></div>'
                f'<div class="score neu"><div class="score-n">{cnt.get("Neutral",0)}</div>'
                '<div class="score-l">Neutral</div></div>'
                f'<div class="score neg"><div class="score-n">{cnt.get("No",0)}</div>'
                '<div class="score-l">Negative</div></div>'
                f'<div class="score nod"><div class="score-n">{cnt.get("No Data",0)}</div>'
                '<div class="score-l">No data</div></div>'
                '</div>', unsafe_allow_html=True)
            st.caption("Grouped by theme — expand a category to see its checks.")
            by_name = {c["name"]: c for c in checks}
            for cat, names in HEALTH_GROUPS:
                gc = [by_name[n] for n in names if n in by_name]
                if not gc:
                    continue
                pos = sum(1 for c in gc if c["verdict"] == "Yes")
                with st.expander(f"{cat}  ·  {pos}/{len(gc)} positive", expanded=False):
                    st.markdown("".join(
                        f'<div class="chk"><span class="tag {c["verdict"].replace(" ","")}">'
                        f'{c["verdict"]}</span><span class="chk-n">{c["name"]}</span>'
                        f'<span class="chk-d">{c["detail"]}</span></div>' for c in gc),
                        unsafe_allow_html=True)
            st.caption("Mechanical checks on reported annual financials. A negative flag "
                       "marks something worth investigating, not a verdict on the company. "
                       "Ratios like debt-to-equity do not carry the same meaning for banks "
                       "and NBFCs. Not covered: promoter pledging, auditor changes, "
                       "related-party transactions, contingent liabilities.")


@st.dialog("Stock detail", width="small")
def detail_dialog(symbol: str):
    render_detail(symbol)
    # make the modal draggable (by its top strip) and resizable
    components.html('''
<script>
(function(){
  var d = window.parent.document;
  function init(){
    var dlg = d.querySelector('div[role="dialog"]');
    if(!dlg){ return setTimeout(init, 120); }
    if(dlg.dataset.mv){ return; }
    dlg.dataset.mv = "1";
    dlg.style.resize = "both"; dlg.style.overflow = "auto"; dlg.style.position = "fixed";
    var sx, sy, ox, oy, drag = false;
    dlg.addEventListener("mousedown", function(e){
      if(e.target.closest('button,input,a,select,textarea,[role="tab"]')) return;
      var r = dlg.getBoundingClientRect();
      if(e.clientY - r.top > 56) return;   // only the top strip is a drag handle
      drag = true; sx = e.clientX; sy = e.clientY; ox = r.left; oy = r.top;
      dlg.style.left = ox+"px"; dlg.style.top = oy+"px"; dlg.style.margin = "0";
      dlg.style.cursor = "grabbing"; e.preventDefault();
    });
    d.addEventListener("mousemove", function(e){
      if(!drag) return;
      dlg.style.left = (ox + e.clientX - sx)+"px";
      dlg.style.top  = (oy + e.clientY - sy)+"px";
    });
    d.addEventListener("mouseup", function(){ drag = false; dlg.style.cursor = "default"; });
  }
  init();
})();
</script>
''', height=0)


# ------------------------------ visitor stats -----------------------------
_STATS_FILE = "visitor_stats.json"


def _inject_ga() -> None:
    """Load Google Analytics 4 into the top-level page. Free, durable, and the
    dashboard (unique users, live users, geography) is private to your GA login.
    Set GA_MEASUREMENT_ID = 'G-XXXXXXX' in app secrets to enable."""
    gid = "G-FL4MEDNP7H"
    try:
        gid = str(st.secrets.get("GA_MEASUREMENT_ID", "")).strip() or gid
    except Exception:
        pass
    if not gid:
        return
    components.html(f"""
        <!-- Google tag (gtag.js) -->
        <script async src="https://www.googletagmanager.com/gtag/js?id={gid}"></script>
        <script>
          window.dataLayer = window.dataLayer || [];
          function gtag(){{ dataLayer.push(arguments); }}
          gtag('js', new Date());
          var ref = document.referrer || 'https://stockmerit.streamlit.app';
          gtag('config', '{gid}', {{
            page_location: ref,
            page_referrer: ref
          }});
        </script>
    """, height=0)


def _track_and_get_stats() -> dict:
    """Count unique browsers (via a first-party cookie) and total page hits.
    Stored in a local JSON file. Ephemeral on Streamlit Cloud - see notes."""
    try:
        with open(_STATS_FILE) as fh:
            data = json.load(fh)
    except Exception:
        data = {"unique": [], "hits": 0}
    vid = None
    try:
        vid = st.context.cookies.get("sm_vid")
    except Exception:
        pass
    if not vid:
        vid = uuid.uuid4().hex
        components.html(
            f"<script>document.cookie='sm_vid={vid};max-age=31536000;path=/;SameSite=Lax';</script>",
            height=0)
    if vid not in data["unique"]:
        data["unique"].append(vid)
    data["hits"] = int(data.get("hits", 0)) + 1
    try:
        with open(_STATS_FILE, "w") as fh:
            json.dump(data, fh)
    except Exception:
        pass
    return data


def _maybe_show_admin_stats() -> None:
    try:
        key = str(st.secrets.get("ADMIN_KEY", "")).strip()
    except Exception:
        key = ""
    if not key or str(st.query_params.get("admin", "")) != key:
        return
    try:
        with open(_STATS_FILE) as fh:
            d = json.load(fh)
    except Exception:
        d = {"unique": [], "hits": 0}
    c1, c2 = st.columns(2)
    c1.metric("Unique visitors", f"{len(d.get('unique', [])):,}")
    c2.metric("Total page hits", f"{int(d.get('hits', 0)):,}")
    st.caption("Visible only with your private ?admin key. Counts reset if the "
               "app is redeployed (ephemeral storage).")


# ---------------------------------- app -----------------------------------

st.set_page_config(page_title="StockMerit — NSE RSI Screener",
                   page_icon="📊", layout="wide",
                   initial_sidebar_state="expanded")
st.markdown(CSS, unsafe_allow_html=True)

_inject_ga()
_track_and_get_stats()
_maybe_show_admin_stats()


def _valid_symbol(sym: str) -> bool:
    """Whitelist symbol input to block injection via search box / URL params."""
    return bool(re.fullmatch(r"[A-Za-z0-9&.\-]{1,20}", sym or ""))


_BAND_HTML = (
    '<div class="band"><div class="band-left"><div class="band-name">'
    '<span class="brandmark"><svg width="22" height="22" viewBox="0 0 32 32" fill="none">'
    '<defs><linearGradient id="smg" x1="2" y1="26" x2="30" y2="6" gradientUnits="userSpaceOnUse">'
    '<stop stop-color="#1B4DB8"/><stop offset="1" stop-color="#3E9BFF"/></linearGradient></defs>'
    '<path d="M4 22 L13 15 L18 19 L27 8" stroke="url(#smg)" stroke-width="3.4" '
    'stroke-linecap="round" stroke-linejoin="round"/>'
    '<path d="M20 8 H28 V16" stroke="url(#smg)" stroke-width="3.4" '
    'stroke-linecap="round" stroke-linejoin="round"/></svg></span>'
    '<span style="letter-spacing:-0.02em"><span style="color:#FFFFFF">Stock</span>'
    '<span style="color:#5CA8FF">Mer</span>'
    '<span style="color:#FFFFFF">it</span></span></div>'
    '<div class="band-sub">Analyze it. Stock it.</div></div>'
    f'<div class="band-date">{dt.date.today().strftime("%d-%m-%Y")}</div></div>')


def _open_in_screener(sym: str) -> None:
    """Show a searched stock as a one-row screener result. The user then clicks
    the symbol to open its detail page (the regular flow)."""
    _sy = sym.strip().upper()
    if not _valid_symbol(_sy):
        return
    st.session_state["results"] = scan((f"{_sy}.NS",), 0.0, 1)
    st.session_state["scanned"] = 1
    st.session_state["threshold"] = 0
    st.session_state["view"] = "Screener"
    st.session_state.pop("qp_opened", None)
    if "stock" in st.query_params:
        del st.query_params["stock"]
    if "scan" in st.query_params:
        del st.query_params["scan"]
    st.rerun()


# --- live-tick settings (kept in code; no sidebar UI) ---
live_on = True
live_every = 1

# --- top nav: Screener / Stock OI / News ---
_view_map = {"screener": "Screener", "custom": "Custom Screen",
             "mf": "Mutual Funds", "etf": "ETFs",
             "oi": "Stock OI", "news": "News"}
_qp_view = str(st.query_params.get("view", "")).lower()
if _qp_view in _view_map and "view" not in st.session_state:
    st.session_state["view"] = _view_map[_qp_view]
st.session_state.setdefault("view", "Screener")
st.sidebar.markdown('<div class="nav-h">Sections</div>', unsafe_allow_html=True)
view = st.sidebar.radio("view", ["Screener", "Stock OI", "Custom Screen", "ETFs",
                                "Mutual Funds", "News"],
                        label_visibility="collapsed", key="view")
st.sidebar.markdown('<div class="nav-foot">Collapse this panel with the arrow above. '
                    'Data is reference only — not investment advice.</div>',
                    unsafe_allow_html=True)

# --- header row: compact banner, with the stock search beside it on the tabs
# where a stock lookup makes sense (mutual funds and ETFs have their own search)
_SEARCH_TABS = {"Screener", "Custom Screen", "Stock OI", "News"}
if view in _SEARCH_TABS:
    _hb, _sbox = st.columns([3, 1], gap="medium")
    _hb.markdown(_BAND_HTML, unsafe_allow_html=True)
    _sbox.markdown('<div class="hdr-search"></div>', unsafe_allow_html=True)
    with _sbox:
        _all_syms = tickers_for("All NIFTY Stocks")
        if _all_syms:
            _opts = [s.replace(".NS", "") for s in _all_syms]
            _picked = st.selectbox(" ", _opts, index=None,
                                   placeholder="🔍  Search any stock…",
                                   label_visibility="collapsed")
            if _picked and st.session_state.get("last_search") != _picked:
                st.session_state["last_search"] = _picked
                _open_in_screener(_picked)
        else:
            _typed = st.text_input(" ", placeholder="🔍  Search any stock…",
                                   label_visibility="collapsed")
            _tv = _typed.strip().upper()
            if _tv and _valid_symbol(_tv) and st.session_state.get("last_search") != _tv:
                st.session_state["last_search"] = _tv
                _open_in_screener(_tv)
else:
    st.markdown(_BAND_HTML, unsafe_allow_html=True)


# --- hyperlink handler: ?stock=SYMBOL opens the detail dialog in any view ---
qp_stock = st.query_params.get("stock")
if qp_stock and _valid_symbol(str(qp_stock)) and st.session_state.get("qp_opened") != qp_stock:
    st.session_state["qp_opened"] = qp_stock
    detail_dialog(str(qp_stock).upper())

if view == "News":
    st.markdown("### Financial news")
    _news = fetch_news(14)
    if _news:
        for _n in _news:
            st.markdown(
                f'<a href="{_n["url"]}" target="_blank" style="font-size:0.95rem; '
                f'font-weight:600; line-height:1.35; display:block;">{_n["title"]}</a>'
                f'<div style="font-size:0.74rem; color:#0E7C86; margin:2px 0 12px;">'
                f'{_n["source"]}</div>', unsafe_allow_html=True)
        st.caption("Headlines from public RSS feeds — click to read at the source.")
    else:
        st.info("News feed unavailable right now.")
    st.stop()

if view == "Stock OI":
    _oi_rows_data, _oi_src = get_oi_buildup()
    st.markdown("### Open interest — buildup")
    _now_ist = (dt.datetime.utcnow() + dt.timedelta(hours=5, minutes=30)).strftime("%d %b %Y, %H:%M IST")
    if _oi_src == "dhan":
        _cap = f"Live prices &amp; OI via Dhan API - {_now_ist}. "
    elif _oi_src == "yf":
        _cap = (f"Live prices via public market feed - {_now_ist}. "
                "OI needs a Dhan token in app secrets (shown as - until then). ")
    else:
        _cap = "Live market feed is unreachable right now - no data shown. "
    st.caption(_cap + "Rise in OI with rise in price = long buildup.")
    _oi_head = "".join(f'<th class="th-{a}">{h}</th>' for h, a in [
        ("Stock Symbol", "l"), ("LTP", "r"), ("Rise In Price", "r"),
        ("Open Interest", "r"), ("Rise In OI", "r"), ("Signal", "c")])
    _oi_rows = []
    for _r in _oi_rows_data:
        _pc, _oc = _r["price_chg"], _r["oi_chg"]
        _sig, _sc = oi_signal(_pc, _oc)
        _pcl = "#0B7A4B" if _pc >= 0 else "#B3261E"
        _ocl = "#0B7A4B" if (_oc is not None and _oc >= 0) else "#B3261E"
        _oi_txt = f'{int(_r["oi"]):,}' if _r["oi"] is not None else "-"
        _oc_txt = f'{_oc:+.2f}%' if _oc is not None else "-"
        _oi_rows.append(
            "<tr>"
            f'<td><a class="c-sym" href="?view=oi&stock={_r["sym"]}" target="_self">{_r["sym"]}</a></td>'
            f'<td class="c-num">{_r["ltp"]:,.2f}</td>'
            f'<td class="c-num" style="color:{_pcl};font-weight:600">{_pc:+.2f}%</td>'
            f'<td class="c-num c-muted">{_oi_txt}</td>'
            f'<td class="c-num" style="color:{_ocl};font-weight:600">{_oc_txt}</td>'
            f'<td class="c-reco"><span class="pill" style="background:{_sc}">{_sig}</span></td>'
            "</tr>")
    st.markdown(
        '<div class="sh-tablewrap"><table class="sh-table"><thead><tr>'
        + _oi_head + "</tr></thead><tbody>" + "".join(_oi_rows)
        + "</tbody></table></div>", unsafe_allow_html=True)
    st.stop()

SIGNED_COLS = {"Chg1D", "Chg1W", "Chg1M", "Chg3M", "Chg6M", "Chg1Y", "Vs200",
               "VsVWAP", "Upside", "RevGrowth", "ProfitGrowth", "PremDisc"}
PCT_COLS = {"FromHigh", "FromLow", "ROE", "ROA", "DivYld", "OpMargin",
            "NetMargin", "ATRpct"}


def metric_cell(col: str, v) -> str:
    dp = COL_META.get(col, (col, "T", 2))[2]
    if col in SIGNED_COLS:
        return signed_cell(v, dp)
    if col == "RSI":
        return (f'<td class="c-num c-rsi" style="color:{rsi_color(v)}">'
                f'{fmt_val(v, 1)}</td>')
    if col == "Buy/Sell":
        return (f'<td class="c-reco"><span class="pill" title="{reco_info(v)}" '
                f'style="background:{reco_color(v)}">{_html.escape(str(v))}</span></td>')
    if col == "Sector":
        return (f'<td class="c-sec" style="color:{sector_color(v)}">'
                f'{_html.escape(str(v))}</td>')
    return f'<td class="c-num{" c-muted" if col in ("Volume", "AvgVol20") else ""}">' \
           f'{fmt_val(v, dp, pct=col in PCT_COLS)}</td>'


# ============================ Custom Screen view ===========================
if view == "Custom Screen":
    st.markdown("### Custom screen")
    st.caption("Build your own filter set from every technical and fundamental "
               "value the app holds. Conditions combine with AND. Technicals run "
               f"across the whole list; fundamentals then load for up to "
               f"{CUSTOM_FUND_CAP} technical matches.")

    st.session_state.setdefault("cs_filters", [])
    st.session_state.setdefault("cs_universe", "NIFTY 50")
    st.session_state.setdefault("cs_nonce", 0)
    _nonce = st.session_state["cs_nonce"]

    st.markdown('<div class="sec-label">Stocks to scan</div>', unsafe_allow_html=True)
    _urow = st.columns([2, 2, 2, 1, 1], gap="small")
    for _c, _n in zip(_urow[:3], UNIVERSES):
        if _c.button(_n, use_container_width=True, key=f"cs_u_{_n}",
                     type="primary" if st.session_state["cs_universe"] == _n else "secondary"):
            st.session_state["cs_universe"] = _n
            for _k in ("cs_results", "cs_fund"):
                st.session_state.pop(_k, None)
            st.session_state[f"csw_sectors_{_nonce}"] = []
            st.rerun()
    _cs_run = _urow[3].button("Run screen", type="primary", key="cs_run",
                             use_container_width=True)
    _cs_reset = _urow[4].button("Reset", key="cs_reset", use_container_width=True)
    if _cs_reset:
        for _k in [k for k in list(st.session_state)
                   if k.startswith(("cs_", "csw_"))]:
            st.session_state.pop(_k, None)
        # bump the widget nonce so every control below mounts as a NEW widget --
        # popping a key alone is not enough, Streamlit replays the old value from
        # the browser onto a widget that keeps the same key and position
        st.session_state["cs_nonce"] = _nonce + 1
        st.rerun()

    st.markdown('<div class="sec-label" style="margin-top:12px">Add a condition</div>',
                unsafe_allow_html=True)
    _opts = [f"{'Technical' if g == 'T' else 'Fundamental'} · {lab}"
             for _c2, lab, g, _d in NUM_COLS]
    _by_label = {f"{'Technical' if g == 'T' else 'Fundamental'} · {lab}": col
                 for col, lab, g, _d in NUM_COLS}
    _b = st.columns([3, 1.6, 1.3, 1.3, 1.4], gap="small")
    _m = _b[0].selectbox("Value", _opts, key=f"csw_metric_{_nonce}")
    _op = _b[1].selectbox("Condition", ["at least", "at most", "between"],
                          key=f"csw_op_{_nonce}")
    _v1 = _b[2].number_input("Value", value=None, key=f"csw_v1_{_nonce}",
                             placeholder="number")
    _v2 = _b[3].number_input("Upper", value=None, key=f"csw_v2_{_nonce}",
                             placeholder="number", disabled=_op != "between")
    _b[4].markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
    if _b[4].button("Add", use_container_width=True, key="cs_add"):
        if _v1 is None or (_op == "between" and _v2 is None):
            st.warning("Enter a value for the condition.")
        else:
            st.session_state["cs_filters"].append(
                {"col": _by_label[_m], "op": _op, "v1": float(_v1),
                 "v2": float(_v2) if _v2 is not None else None})
            st.rerun()

    _pc = st.columns([3, 3, 1.4], gap="small")
    _preset = _pc[0].selectbox("Preset", ["Start from a preset..."] + list(PRESETS),
                              key=f"csw_preset_{_nonce}")
    _sectors = _pc[1].multiselect(
        "Sector", sorted({"Financial Services", "Technology", "Healthcare",
                          "Consumer Cyclical", "Consumer Defensive", "Industrials",
                          "Basic Materials", "Energy", "Utilities",
                          "Communication Services", "Real Estate"}),
        key=f"csw_sectors_{_nonce}", placeholder="All sectors")
    _pc[2].markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
    if _pc[2].button("Apply preset", use_container_width=True, key="cs_apply_preset"):
        if _preset in PRESETS:
            st.session_state["cs_filters"] = [
                {"col": c, "op": o, "v1": float(a), "v2": float(b) if b is not None else None}
                for c, o, a, b in PRESETS[_preset]]
            st.rerun()

    if st.session_state["cs_filters"]:
        st.markdown('<div class="sec-label" style="margin-top:10px">Active conditions</div>',
                    unsafe_allow_html=True)
        for _i, _f in enumerate(list(st.session_state["cs_filters"])):
            _lab, _grp, _dp = COL_META[_f["col"]]
            _txt = (f"{_lab} between {_f['v1']:g} and {_f['v2']:g}"
                    if _f["op"] == "between" else f"{_lab} {_f['op']} {_f['v1']:g}")
            _fr = st.columns([6, 1], gap="small")
            _fr[0].markdown(
                f'<div style="padding:6px 0;font-size:0.9rem">'
                f'<span class="pill" style="background:'
                f'{"#0E7C86" if _grp == "T" else "#6B5BC7"}">'
                f'{"TECH" if _grp == "T" else "FUND"}</span>&nbsp; {_html.escape(_txt)}</div>',
                unsafe_allow_html=True)
            if _fr[1].button("Remove", key=f"cs_rm_{_i}", use_container_width=True):
                st.session_state["cs_filters"].pop(_i)
                st.rerun()
    else:
        st.caption("No conditions yet — a screen with none returns the whole list, "
                   "ranked by whichever column you sort on.")

    if _cs_run:
        _tks = tickers_for(st.session_state["cs_universe"])
        if not _tks:
            st.error(f"The {st.session_state['cs_universe']} list did not load. "
                     "Try again in a minute.")
        else:
            _df, _priced, _hasf = custom_screen(
                tuple(_tks), st.session_state["cs_filters"], _sectors)
            st.session_state["cs_results"] = _df
            st.session_state["cs_priced"] = _priced
            st.session_state["cs_fund"] = _hasf

    _res = st.session_state.get("cs_results")
    if _res is None:
        st.info("Add conditions (or apply a preset), pick a list, then run the screen.")
    elif _res.empty:
        st.warning("No stock matches every condition. Loosen one and run again.")
    else:
        _hasf = st.session_state.get("cs_fund", False)
        _cols = ["Symbol"] + (["Sector"] if _hasf else []) + ["Price", "Chg1D", "RSI"]
        for _f in st.session_state["cs_filters"]:
            if _f["col"] not in _cols:
                _cols.append(_f["col"])
        if _hasf and "Buy/Sell" not in _cols:
            _cols.append("Buy/Sell")
        _cols = [c for c in _cols[:13] if c in _res.columns]

        _sortable = [c for c in _cols if c not in ("Symbol", "Sector", "Buy/Sell")]
        _sc = st.columns([2, 2, 2.4, 1.6], gap="small")
        _sort = _sc[0].selectbox("Sort by", _sortable,
                                 format_func=lambda c: COL_META[c][0],
                                 key=f"csw_sort_{_nonce}")
        _dir = _sc[1].selectbox("Order", ["High to low", "Low to high"],
                                key=f"csw_dir_{_nonce}")
        _view_df = _res.sort_values(_sort, ascending=_dir == "Low to high",
                                    na_position="last")
        _sc[2].markdown(f"**{len(_res)} matches** from "
                        f"{st.session_state.get('cs_priced', 0)} priced securities.")
        _sc[3].download_button(
            "⤓  Download", key="cs_dl",
            data=df_to_excel_bytes(_view_df, "Custom screen",
                                   {c: COL_META[c][0] for c in COL_META}),
            file_name=f"Custom_Screen_{dt.date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

        _heads = [("Symbol", "l")] + [
            (COL_META[c][0] if c in COL_META else c,
             "l" if c in ("Sector",) else "c" if c == "Buy/Sell" else "r")
            for c in _cols[1:]]
        _rows = []
        for _, _r2 in _view_df.iterrows():
            _cells = [f'<td><a class="c-sym" href="?view=custom&stock={_r2["Symbol"]}" '
                      f'target="_self">{_html.escape(str(_r2["Symbol"]))}</a></td>']
            _cells += [metric_cell(c, _r2.get(c)) for c in _cols[1:]]
            _rows.append("<tr>" + "".join(_cells) + "</tr>")
        st.markdown('<div class="sh-hint">Click a symbol for the full detail view. '
                    "Columns follow your conditions.</div>", unsafe_allow_html=True)
        st.markdown(sh_table(_heads, _rows), unsafe_allow_html=True)
        if not _hasf:
            st.caption("Fundamental values load only when a fundamental condition or "
                       "a sector filter is part of the screen.")

    st.markdown(f'<div class="disc"><strong>Disclaimer</strong> — {DISCLAIMER}</div>',
                unsafe_allow_html=True)
    st.stop()

# ============================= Mutual Funds view ===========================
if view == "Mutual Funds":
    st.markdown("### Mutual funds")
    _mf = load_mutual_funds()
    if _mf.empty:
        st.error("AMFI's NAV file is unreachable right now. Nothing is shown rather "
                 "than stale values — try again in a few minutes.")
        st.stop()
    st.caption(f"{len(_mf):,} schemes from AMFI's official daily NAV file "
               f"(latest published {_mf['Date'].iloc[0]}). Returns come from each "
               "scheme's own NAV history; over 1 year they are annualised.")

    _f1 = st.columns([2, 2, 2.4, 1.6], gap="small")
    _type = _f1[0].selectbox("Type", ["All types"] + sorted(_mf["Type"].unique()),
                             key="mf_type")
    _pool = _mf if _type == "All types" else _mf[_mf["Type"] == _type]
    _amc = _f1[1].selectbox("Fund house", ["All fund houses"] + sorted(_pool["AMC"].unique()))
    if _amc != "All fund houses":
        _pool = _pool[_pool["AMC"] == _amc]
    _cat = _f1[2].selectbox("Category", ["All categories"] + sorted(_pool["Category"].unique()))
    if _cat != "All categories":
        _pool = _pool[_pool["Category"] == _cat]
    _q = _f1[3].text_input("Search scheme", key="mf_q", placeholder="e.g. flexi cap")
    if _q and _q.strip():
        _pool = _pool[_pool["Scheme"].str.contains(_q.strip(), case=False, na=False)]

    _o1 = st.columns([2, 2, 2.4, 1.6], gap="small")
    _plan = _o1[0].selectbox("Plan", ["Any plan", "Direct only", "Regular only"],
                             key="mf_plan")
    if _plan == "Direct only":
        _pool = _pool[_pool["Scheme"].str.contains("direct", case=False, na=False)]
    elif _plan == "Regular only":
        _pool = _pool[~_pool["Scheme"].str.contains("direct", case=False, na=False)]
    _growth = _o1[1].selectbox("Option", ["Any option", "Growth only"], key="mf_opt")
    if _growth == "Growth only":
        _pool = _pool[_pool["Scheme"].str.contains("growth", case=False, na=False)]
    _pool = _pool.reset_index(drop=True)
    _shown = _pool.head(300)
    _o1[2].markdown(f"**{len(_pool):,} schemes** match" +
                    (f" — showing the first {len(_shown)}." if len(_pool) > 300 else "."))
    _o1[3].download_button(
        "⤓  Download", key="mf_dl",
        data=df_to_excel_bytes(_pool.drop(columns=["Code"]), "Mutual funds"),
        file_name=f"Mutual_Funds_{dt.date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)

    st.session_state.setdefault("mf_ret", {})
    _rc = st.columns([2, 4], gap="small")
    _n_ret = min(len(_shown), 25)
    if _rc[0].button(f"Load returns for top {_n_ret}", key="mf_load_ret",
                     use_container_width=True, disabled=_n_ret == 0):
        _bar = st.progress(0.0, text="Reading NAV histories...")
        for _i, _code in enumerate(_shown["Code"].head(_n_ret).tolist()):
            st.session_state["mf_ret"][_code] = mf_returns(_code)
            _bar.progress((_i + 1) / _n_ret, text=f"Loaded {_i + 1}/{_n_ret}")
        _bar.empty()
    _rc[1].caption("Returns are fetched per scheme, so they load on demand for the "
                   "visible top of the list.")

    _mf_heads = [("Scheme", "l"), ("Fund house", "l"), ("Category", "l"),
                 ("NAV (Rs)", "r"), ("1Y %", "r"), ("3Y %", "r"), ("5Y %", "r"),
                 ("NAV date", "l")]
    _mf_rows = []
    for _, _s in _shown.iterrows():
        _ret = st.session_state["mf_ret"].get(_s["Code"], {})
        _mf_rows.append(
            "<tr>"
            f'<td style="white-space:normal;max-width:420px;font-weight:500">'
            f'{_html.escape(str(_s["Scheme"]))}</td>'
            f'<td class="c-sec" style="color:{sector_color(_s["AMC"])}">'
            f'{_html.escape(str(_s["AMC"]))}</td>'
            f'<td class="c-muted" style="white-space:normal;max-width:260px">'
            f'{_html.escape(str(_s["Category"]))}</td>'
            f'<td class="c-num">{fmt_val(_s["NAV"], 4)}</td>'
            + signed_cell(_ret.get("1Y")) + signed_cell(_ret.get("3Y"))
            + signed_cell(_ret.get("5Y"))
            + f'<td class="c-date">{_html.escape(str(_s["Date"]))}</td>'
            "</tr>")
    st.markdown(sh_table(_mf_heads, _mf_rows), unsafe_allow_html=True)
    st.markdown(f'<div class="disc"><strong>Disclaimer</strong> — {DISCLAIMER}</div>',
                unsafe_allow_html=True)
    st.stop()

# ================================ ETFs view ================================
if view == "ETFs":
    st.markdown("### ETFs")
    _etf_live = market_open()

    @st.fragment(run_every=(60 if _etf_live else None))
    def render_etf_tab():
        # prices refresh on a 1-minute bucket, iNAVs on a 5-minute bucket, so
        # every figure on this tab tracks live trade while the market is open
        _etf, _nse_live = scan_etfs(int(time.time() // 60) if _etf_live else 0,
                                    int(time.time() // 300) if _etf_live else 0)
        if _etf.empty:
            st.error("The market feed is unreachable right now — no ETF data shown.")
            return
        _ts = dt.datetime.now(IST).strftime("%H:%M:%S")
        _src = ("prices and iNAV from the NSE ETF board" if _nse_live
                else "prices live; iNAV from the fund quote feed")
        st.caption(
            (f"🟢 Live · {len(_etf)} NSE-listed ETFs · {_src}, updated {_ts} IST "
             "and refreshing every minute."
             if _etf_live else
             f"⚪ Market closed · {len(_etf)} NSE-listed ETFs at last close ({_ts} IST).")
            + " Click a symbol for its price levels.")

        _e1 = st.columns([2, 2, 2.4, 1.6], gap="small")
        _ecat = _e1[0].selectbox("Category", ["All categories"] + ETF_CATEGORIES, key="etf_cat")
        _pool = _etf if _ecat == "All categories" else _etf[_etf["Category"] == _ecat]
        _eq = _e1[1].text_input("Search", key="etf_q", placeholder="e.g. gold",
                                label_visibility="visible")
        if _eq and _eq.strip():
            _m = _eq.strip()
            _pool = _pool[_pool["Symbol"].str.contains(_m, case=False, na=False)
                          | _pool["Name"].str.contains(_m, case=False, na=False)]
        _esort = _e1[2].selectbox(
            "Sort by", ["Chg1D", "Chg1M", "Chg1Y", "RSI", "PremDisc", "Volume",
                        "Price", "Vs200"],
            format_func=lambda c: COL_META[c][0], key="etf_sort")
        _pool = _pool.sort_values(_esort, ascending=False, na_position="last")
        _e1[3].download_button(
            "⤓  Download", key="etf_dl",
            data=df_to_excel_bytes(
                _pool[["Symbol", "Name", "Category", "Price", "iNAV", "PremDisc",
                       "Chg1D", "Chg1M", "Chg1Y", "RSI", "Vs200", "High52", "Low52",
                       "Volume", "AvgVol20"]],
                "ETFs", {c: COL_META[c][0] for c in COL_META}),
            file_name=f"ETFs_{dt.date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

        _e_cols = ["Price", "iNAV", "PremDisc", "Chg1D", "Chg1W", "Chg1M", "Chg1Y",
                   "RSI", "Vs200", "High52", "Low52", "Volume"]
        _e_heads = [("Symbol", "l"), ("ETF", "l"), ("Category", "l")] + \
                   [(COL_META[c][0], "r") for c in _e_cols]
        _e_rows = []
        for _, _r3 in _pool.iterrows():
            _e_rows.append(
                "<tr>"
                f'<td><a class="c-sym" href="?view=etf&stock={_r3["Symbol"]}" '
                f'target="_self">{_html.escape(str(_r3["Symbol"]))}</a></td>'
                f'<td style="font-weight:500">{_html.escape(str(_r3["Name"]))}</td>'
                f'<td class="c-sec" style="color:{sector_color(_r3["Category"])}">'
                f'{_html.escape(str(_r3["Category"]))}</td>'
                + "".join(metric_cell(c, _r3.get(c)) for c in _e_cols)
                + "</tr>")
        st.markdown(sh_table(_e_heads, _e_rows), unsafe_allow_html=True)
        st.caption("iNAV is the fund's indicative net asset value per unit, refreshed "
                   "with the prices; prem / disc is how far the traded price sits "
                   "above or below it. Gold, silver and international ETFs track their "
                   "own underlying, so RSI and moving averages describe the ETF price, "
                   "not an index.")

    render_etf_tab()
    st.markdown(f'<div class="disc"><strong>Disclaimer</strong> — {DISCLAIMER}</div>',
                unsafe_allow_html=True)
    st.stop()

# --- Screener view ---
_live_active = live_on and market_open()

@st.fragment(run_every=(live_every if _live_active else None))
def render_indices_strip():
    bucket = int(time.time() // max(int(live_every), 5))
    _idx = fetch_indices(bucket)
    if _idx:
        for _c, _ix in zip(st.columns(len(_idx)), _idx):
            _clr = "#0B7A4B" if _ix["chg"] >= 0 else "#B3261E"
            _arr = "▲ +" if _ix["chg"] >= 0 else "▼ "
            _c.markdown(
                f'<div class="idxcard"><div class="idx-n">{_ix["name"]}</div>'
                f'<div class="idx-v">{_ix["value"]:,.2f}</div>'
                f'<div class="idx-c" style="color:{_clr}">{_arr}{abs(_ix["chg"]):.2f}%</div></div>',
                unsafe_allow_html=True)
    _ts = dt.datetime.now(IST).strftime("%H:%M:%S")
    if _live_active:
        st.caption(f"🟢 Live · updated {_ts} IST · refreshing every {int(live_every)}s")
    elif live_on:
        st.caption(f"⚪ Market closed · showing last close ({_ts} IST)")
    else:
        st.caption(f"⏸ Live updates paused · {_ts} IST")

render_indices_strip()

# --- universe + actions on one line ---
st.markdown('<div class="sec-label">Stocks to scan</div>', unsafe_allow_html=True)
if "universe" not in st.session_state:
    st.session_state["universe"] = "NIFTY 50"
st.session_state.setdefault("rsi_thr", 65)
row = st.columns([2, 2, 2, 1, 1], gap="small")
for col, name in zip(row[:3], UNIVERSES):
    if col.button(name, use_container_width=True,
                  type="primary" if st.session_state["universe"] == name else "secondary",
                  key=f"u_{name}"):
        st.session_state["universe"] = name
        st.session_state.pop("bs_filter", None)
        st.rerun()
universe_choice = st.session_state["universe"]
uploaded = None
run = row[3].button("Run", type="primary", use_container_width=True)
clear = row[4].button("Clear", use_container_width=True)

st.markdown('<div class="sec-label" style="margin-top:12px">Filters</div>',
            unsafe_allow_html=True)
fcol, _fsp = st.columns([3, 2], gap="large")
threshold = fcol.slider(f"RSI at or above — {st.session_state['rsi_thr']}", 0, 90,
                        key="rsi_thr")

if universe_choice == "All NIFTY Stocks":
    st.caption("All NIFTY Stocks pulls ~2000 listed stocks and can take several "
               "minutes. NIFTY 500 is the better broad option for daily use.")

if clear:
    for k in ("results", "scanned", "threshold", "opened_for", "screener_table",
              "qp_opened", "rsi_thr", "universe", "view", "bs_filter"):
        st.session_state.pop(k, None)
    st.query_params.clear()
    st.rerun()

if run:
    tickers = tickers_for(universe_choice, uploaded)
    if not tickers:
        st.error(f"The {universe_choice} list did not load. Try again in a minute.")
        st.stop()
    st.session_state["results"] = scan(tuple(tickers), float(threshold), FUND_LIMIT)
    st.session_state["scanned"] = len(tickers)
    st.session_state["threshold"] = threshold
    st.session_state["opened_for"] = None
    st.session_state.pop("bs_filter", None)
    st.query_params["scan"] = _cfg_str(universe_choice, threshold, FUND_LIMIT)
    if "stock" in st.query_params:
        del st.query_params["stock"]
    st.session_state.pop("qp_opened", None)

# restore the results table after a ?stock= hyperlink reload (cached -> instant)
if st.session_state.get("results") is None:
    _cfg = st.query_params.get("scan")
    if _cfg:
        try:
            _u, _th, _n = _cfg.split(SCAN_SEP)
            _th, _n = float(_th), int(_n)
            _tks = tickers_for(_u)
            if _tks:
                st.session_state["results"] = scan(tuple(_tks), _th, _n)
                st.session_state["scanned"] = len(_tks)
                st.session_state["threshold"] = _th
                st.session_state["universe"] = _u
        except Exception:
            pass

results = st.session_state.get("results")

if results is None:
    st.info("Pick a stock list, set your RSI threshold, then run the screen. "
            "You can also look up any single stock from the search box above.")
elif results.empty:
    st.warning(f"Nothing is at RSI {st.session_state.get('threshold', threshold)} or "
               "above right now. Lower the threshold or widen the stock list.")
else:
    _all_bs = [b for b in ["Buy", "Sell", "Hold", "Neutral"]
               if b in set(results["Buy/Sell"].astype(str))]
    _fc1, _fc2 = st.columns([2, 6], gap="small")
    _bs_choice = _fc1.selectbox("Buy/Sell", ["All"] + _all_bs, index=0, key="bs_filter")
    if _bs_choice != "All":
        results = results[results["Buy/Sell"].astype(str) == _bs_choice]
    hleft, hright = st.columns([4, 1], gap="small")
    hleft.markdown(f"**{len(results)} stocks** at RSI "
                   f"{st.session_state.get('threshold', threshold)} or above, "
                   f"from {st.session_state.get('scanned', 0)} scanned.")
    hright.download_button("⤓  Download report", data=to_excel_bytes(results),
                           file_name=f"RSI_Screener_{dt.date.today().isoformat()}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet", use_container_width=True)

    def _fmt(x):
        return f"{x:,.2f}" if isinstance(x, (int, float)) else str(x)

    aligns = ["l", "l", "l", "r", "r", "r", "c", "r", "r", "c", "r", "r"]
    head_html = "".join(
        f'<th class="th-{a}">{COL_LABELS.get(h, h)}</th>' for h, a in zip(HEADERS, aligns))

    _cfgq = quote(st.query_params.get("scan", ""), safe="")
    body_rows = []
    for _, r in results.iterrows():
        sym = r["Stock Symbol"]
        rsi = r["RSI"]
        reco = r["Buy/Sell"]
        rsi_txt = f"{rsi:.1f}" if isinstance(rsi, (int, float)) else str(rsi)
        href = f"?scan={_cfgq}&stock={sym}" if _cfgq else f"?stock={sym}"
        _sv = lambda v: f"{v:,.2f}" if isinstance(v, (int, float)) else "n/a"
        _vol = lambda v: f"{int(v):,}" if isinstance(v, (int, float)) else "n/a"
        sma_tip = (f"20-day SMA: {_sv(r.get('_SMA20'))}  |  "
                   f"50-day SMA: {_sv(r.get('_SMA50'))}  |  "
                   f"200-day SMA: {_sv(r.get('_SMA200'))}")
        _pe = _sv(r.get("PE"))
        _secpe = _sv(r.get("Sec PE"))
        _vw_tip = ("VWAP = SUM(close x volume) / SUM(volume) over the last 20 sessions. "
                   f'Closing price is {r["VWAP"]} the 20-day VWAP.')
        body_rows.append(
            "<tr>"
            f'<td class="c-date">{r["Date"]}</td>'
            f'<td><a class="c-sym" href="{href}" target="_self">{sym}</a></td>'
            f'<td class="c-sec" style="color:{sector_color(r["Sector"])}">{_html.escape(str(r["Sector"]))}</td>'
            f'<td class="c-num c-price" title="{sma_tip}">{_fmt(r["Current Price (Rs)"])}</td>'
            f'<td class="c-num c-muted">{_vol(r.get("Volume"))}</td>'
            f'<td class="c-num c-rsi" style="color:{rsi_color(rsi)}">{rsi_txt}</td>'
            f'<td class="c-vwap" title="{_vw_tip}" style="text-align:center;font-weight:600;'
            f'color:{"#0B7A4B" if r["VWAP"]=="Above" else "#B3261E" if r["VWAP"]=="Below" else "#5E6E7E"}">'
            f'{r["VWAP"]}</td>'
            f'<td class="c-num">{_pe}</td>'
            f'<td class="c-num c-muted">{_secpe}</td>'
            f'<td class="c-reco"><span class="pill" title="{reco_info(reco)}" '
            f'style="background:{reco_color(reco)}">{reco}</span></td>'
            f'<td class="c-num c-muted">{_fmt(r["52 Week High (Rs)"])}</td>'
            f'<td class="c-num">{_fmt(r["1 Year Target (Rs)"])}</td>'
            "</tr>")

    st.markdown('<div class="sh-hint">Click a stock symbol for its full detail. '
                "Hover the underlined price to see 20 / 50 / 200-day moving averages."
                "</div>", unsafe_allow_html=True)
    st.markdown(
        '<div class="sh-tablewrap"><table class="sh-table"><thead><tr>'
        + head_html + "</tr></thead><tbody>" + "".join(body_rows)
        + "</tbody></table></div>", unsafe_allow_html=True)
    na = int((results["Buy/Sell"] == "n/a").sum())
    if na:
        st.caption(f"{na} of {len(results)} stocks have no published analyst rating — "
                   "common for smaller companies.")

st.markdown(f'<div class="disc"><strong>Disclaimer</strong> — {DISCLAIMER}</div>',
            unsafe_allow_html=True)
